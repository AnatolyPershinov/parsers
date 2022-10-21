#!/usr/bin/env python
# coding: utf-8

# In[ ]:


import openpyxl
import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import pandas as pd

import os
import json
import re
import numpy as np


# In[ ]:


def get_location_by_name(sheet: openpyxl.worksheet.worksheet.Worksheet, target_str: list, 
                         min_row = 1, max_row = None, min_col = 1, max_col = None):
    """Function for finding the position on a page of a cell containing a substring from a list
    Argi:
        sheet (openpyxl.worksheet.worksheet.Worksheet): target sheet in excel file
        target (list): list of substrings
    Returns:
        list: [position_row, position_column, coordinate_of_cell]
    """
    if not max_row:
        max_row = sheet.max_row
    if not max_col:
        max_col = sheet.max_column
        
    target_pos =  []
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            value = sheet.cell(row=row, column = col).value
            coordinate = sheet.cell(row=row, column = col).coordinate
            if value: 
                for name in [var.lower() for var in target_str]:
                    if name == str(value).lower():
                        target_pos.append([row-1, col, coordinate])
    return target_pos


# In[ ]:


def get_width(sheet, coordinate: str):
    """
    Function to find a maximal level of multiindex header across rows
    """
    for MergedCell in sheet.merged_cells.ranges:
        if coordinate in MergedCell:
            x1, x2 = str(MergedCell).split(":")

            return sheet[x2].row - sheet[x1].row + 1
        


# In[ ]:


def get_length(sheet, coordinate: str):
    """
    Function to find a maximal level of multiindex header across columns
    """
    for MergedCell in sheet.merged_cells.ranges:
        if coordinate in MergedCell:
            x1, x2 = str(MergedCell).split(":")

            return sheet[x2].column - sheet[x1].column + 1


# In[ ]:


def get_start_point(sheet, initial_coords:int, excluding:list = [], window:int=10):
    """
    Function to get a coordinate of first not-none row in defining column
    Args:
    sheet
    initial_coords: int, point to start search with
    window: int, range of searching
    structure_shift: int, structure shift due to change of structure (or heurstic)
    excluding: List[str], list of excluding markers
    
    """
    finish = initial_coords[0] + window
    column = sheet[get_column_letter(initial_coords[1])]

    for index, cell in enumerate(column):
        if initial_coords[0] < index <= finish:
            if cell.value and all(f not in str(cell.value) for f in excluding):
                return [cell.row, cell.column, cell.value]


# In[ ]:


def get_sheets(path: str, markers: list):
    """
    Function to get workbooks from sheets
    Args:
    Markers: List[str], engramms searched in names of sheets
    
    """
    wb = load_workbook(path, data_only=True)
    sheetnames = wb.sheetnames
    
    workbooks = []
    for marker in markers:
        if any(marker in sheetname for sheetname in sheetnames):        
            sheet = wb[[sheet for sheet in sheetnames if marker.lower() in sheet.lower()][0]]
            workbooks.append(sheet)
        else:
            print(False)
    return workbooks


# In[ ]:


def get_header(sheet, start, width, fmt="%Y-%m-%d", bounds:list = []):
    """
    start: int, index of row
    width: int, max level of multiindex header
    fmt: str, date format
    bound: List[datetime.datetime], time bounds for header
    """
    merged = sheet.merged_cells.ranges
    header = []
    
    for cell in sheet[start+1]:
        if cell.value:
            if get_width(sheet, cell.coordinate) == width:
                header.append(cell.value)
            else:
                if not isinstance(sheet[cell.coordinate].value, datetime.datetime) and get_length(sheet, cell.coordinate):
                    l = get_length(sheet, cell.coordinate)
                    for i in range(1, l):
                        header.append(" ".join([cell.value, sheet[get_column_letter(cell.column)][cell.row].value]))
                        header.append(" ".join([cell.value, sheet[get_column_letter(cell.column + l - 1)][cell.row].value]))
                elif get_length(sheet, cell.coordinate): # datetime
                    h = get_length(sheet, cell.coordinate)
                    for i in range(h):
                        if not bounds[0] <= sheet[get_column_letter(cell.column + i)][cell.row].value <= bounds[1]:
                            break
                        else:
                            header.append(
                                sheet[get_column_letter(cell.column + i)][cell.row].value.strftime(fmt)
                            )
                else:
                    header.append(cell.value)
        else: # is it None or column without any name?
            if get_width(sheet, cell.coordinate) == width: 
                header.append('unknown' + str(cell.column))
            
    return [i.replace("\n", "") for i in header if type(i) == str]


# In[ ]:


def parse_rows(sheet, start: list, header: list):
    rows = [row for row in sheet.rows][start[0]-1:]
    final = {}
    for row in rows:
        for key, cell in zip(header, row):
            value = cell.value

            if not key in final.keys():
                final[key] = []

            final[key].append(value)
    return final


# In[ ]:


def msg_to_df(sheet_msg, bounds):
    defining_name = 'наименование работ' # the most informative cell
    initial_coordinates = get_location_by_name(sheet_msg, [defining_name])[0]
    
    w = get_width(sheet_msg, initial_coordinates[2])
    
    start = get_start_point(sheet_msg, initial_coordinates)
    
    header = get_header(sheet_msg, start = initial_coordinates[0],
                        width=w, bounds = bounds)
    final = parse_rows(sheet_msg, start, header)
    msg = pd.DataFrame(final)
    msg = msg.dropna(axis=1, how='all')
    msg = msg.drop(["план/факт"], axis=1)    
    return msg


# In[ ]:


def res_to_df(sheet_res, bounds):
    defining_name_resourses = 'ресурсы'
    ic = get_location_by_name(sheet_res, [defining_name_resourses])[0]
    
    header = get_header(sheet_res, start = ic[0], width=get_width(sheet_res, ic[2]), bounds=bounds)
    
    start = get_start_point(sheet_res, ic)
    
    rows = [row for row in sheet_res.rows][start[0]-1:]
    final = {}
    for row in rows:
        for key, cell in zip(header, row):
            value = cell.value

            if not key in final.keys():
                final[key] = []

            final[key].append(value)
    res = pd.DataFrame(final)
    res = res.dropna(axis=1, how='all')
    
    return res   


# In[ ]:


def check_row_onNan(row):
    ONCHECK = ['Всегопо проекту', 'С начала строительства  план', 'С начала строительства  факт']
    if all(pd.isna(a) for a in row[ONCHECK]):
        return True
    else:
        return False


# In[ ]:


def to_date(row: str):
    calendar = {m:n for m,n in zip(["январь", "февраль", "март", "апрель", "май", "июнь", "июль", 
                                    "август", "сентябрь", "октябрь", "ноябрь", "декабрь"], range(1, 13))}
    pattern = "(\w{2,7}).?\s?(\d\d)"
    monat, year = re.findall(pattern, row)[0]
    for month, m in calendar.items():
        if monat in month:
            return datetime.datetime(2000+int(year), m, 1).strftime("%Y.%m.%d")


# In[ ]:


def validate_numeric(data):
    if isinstance(data, str):
        if "%" in data:
            return float(data.replace("%", ""))
        else:
            return np.nan
    else:
        return data


# In[ ]:


def msg_to_json(msg):
    operation_package_start = False
    init_uw = True
    upper_works_start = True
    level = -1
    
    arbeit_index = 0
    
    
    upper_works = []
    parsed = {"work": []}

    for _, mrow in msg.iterrows():
        if not mrow['Наименование работ']:
            if all(pd.isna(a) for a in mrow):
                if operation_package_start:
                    operation_package_start = False
                    level = -1
                else:
                    continue

            else: # fact
                if not operation_package_start:
                    continue

                else: # Fact dates
                    days_data = mrow.dropna()
                    for date, value in zip(days_data.index, days_data):
                        if len(date.split("-")) == 3:
                            for index in range(len(package['work_data']['progress'])):
                                if date in package['work_data']['progress'][index].keys():
                                    package['work_data']['progress'][index][date]['fact'] = value
                    operation_package_start = False
                    level = -1

        else: # info, upper work
            if check_row_onNan(mrow):
                if init_uw:
                    upper_works.append(mrow["Наименование работ"])
#                     print('Up 1 -> ', upper_works)
                    continue
                elif not operation_package_start: # after package start
                    upper_works_start = True
                    try:
                        upper_works[level] = mrow["Наименование работ"]
                    except IndexError:
                        break # in case of huge gap in defining column break the loop
                    if not level == -1: # level shift
                        upper_works.pop(level) 
                        upper_works.append(mrow["Наименование работ"])
        
                    level -= 1                      
            else:          
                upper_works_start = False
                package = {}
                operation_package_start = True
                init_uw = False

#                 print('Up 2 -> ', upper_works)

                ff = upper_works[:]
                package["upper works"] = ff

                package["work title"] = mrow["Наименование работ"]
                package["work id"] = arbeit_index
                package["measurements"] = mrow["Ед.изм"]
                package["amount"] = mrow['Всегопо проекту']
                package["work_data"] = {
                    "start_date": {
                        "plan": to_date(mrow["Начало работ по контракту"]) if mrow["Начало работ по контракту"] else None,
                        "estimate": None,
                        "fact": None
                        },
                    "stop_date": {
                          "plan": to_date(mrow["Окон. работ по контракту"]) if mrow["Окон. работ по контракту"] else None,
                          "estimate": None,
                          "fact": None
                        },
                    "complite_state": {
                          "plan": validate_numeric(mrow['С начала строительства  план']) * 100 / validate_numeric(mrow['Всегопо проекту']),
                          "fact": validate_numeric(mrow['С начала строительства  факт']) * 100 / mrow['Всегопо проекту']
                        },
                    "current_remain": None,
                    "whole_remain": None,
                    "mounth_complite": {
                        "plan": validate_numeric(mrow["Задание на месяц"]) * 100 / validate_numeric(mrow['Всегопо проекту']),
                        "fact": validate_numeric(mrow["С начала месяца факт"]) * 100 / validate_numeric(mrow['Всегопо проекту'])},
                    "progress": []
                }
                for date, value in zip(mrow.index, mrow):
                    if "-" in date:
                        package["work_data"]["progress"].append({date: {"plan": value, "fact": None}})
                parsed['work'].append(package)
                arbeit_index += 1
    return parsed


# In[ ]:


def res_to_json(res):
    res_name = None
    idx = 0
    
    package_start = False
    
    parsed = []
    for _, rrow in res.iterrows(): 
        if rrow['Ресурсы'] == "Итого":
            continue
        if rrow['Ресурсы']:
            if 'ресурсы' in rrow['Ресурсы'].lower():
                res_name = "human" if rrow['Ресурсы'].split()[0].lower() == "людские" else "equipment"
                continue
            
            package_start = True
            rpack = {
              "resource_id": idx,
              "resource_name": rrow['Ресурсы'],
              "type": res_name,
              "progress": []
                }
            parsed.append(rpack) 
            idx+=1
            days_data = rrow.dropna()
            for date, value in zip(rrow.index, rrow):
                if "-" in date:
                    rpack["progress"].append({date: {"plan": value, "fact": None}})
        else:
            if not package_start:
                continue
            for date, value in zip(rrow.index, rrow):
                if "-" in date:
                    for index in range(len(rpack['progress'])):
                        if date in rpack['progress'][index].keys():
                            rpack['progress'][index][date]['fact'] = value
            package_start = False
    return parsed


# In[ ]:


def get_months_range(path):
    calendar = {m:n for m,n in zip(["январь", "февраль", "март", "апрель", "май", "июнь", "июль", 
                                    "август", "сентябрь", "октябрь", "ноябрь", "декабрь"], range(1, 13))}
    info = path.split('\\')[-2]
    month = info.split()[2]
    

    m = calendar[month.lower()]
    
    if m > 12:
        m -= 11
    
    last = datetime.date(2015, m+1 if m < 12 else 1, 1) - datetime.timedelta(days=1)
    last_day = last.day
    
    bounds = [datetime.datetime(2015, m-1, 1), datetime.datetime(2015, m, last_day)]
    
    return bounds


# In[ ]:


PATH = r"C:\Users\Roman\Desktop\Project 234\СМГ УКПГ Н Порт\МСГ УКПГ\МСГ УКПГ"

jahre = [a for a in os.listdir(PATH) if a[0].isdigit()]

monate = []
for jahr in jahre:
    monate.append([f"{jahr}\\{a}" for a in os.listdir(os.path.join(PATH, jahr))])
    
files = {}
for i in monate:
    for monat in i:
        files[os.path.relpath(monat)] = [os.path.join(PATH, monat, file) for file in os.listdir(os.path.join(PATH, monat))]
for d, jahr in files.items():
    if d.endswith(("Октябрь", "декабрь", "май")):
        continue
    else:
        for path in jahr:
            print("\\".join(path.split('\\')[-3:]))
            sheet_msg, sheet_res = get_sheets(path=path, markers=['МСГ', 'Ресурсы'])
            msg = msg_to_df(sheet_msg, bounds=get_months_range(path))
            msg.drop("Накоп. 17.04 - 30.09", axis=1, inplace=True)
#             msg["% вып. плана с начала месяца"] = msg["% вып. плана с начала месяца"] \
#             .apply(lambda x: float(x.replace("%", "")) if isinstance(x, str) else x)
            
            res = res_to_df(sheet_res, bounds=get_months_range(path))
            
            serialized = msg_to_json(msg)
            serialized['resource'] = res_to_json(res)
            serialized['file_name'] = "\\".join(path.split('\\')[-3:])
            name = f"Н.Порт_МСГ_УГПК_" + path.split('\\')[-1][:8]
            with open(f"{name}.json", 'w', encoding='utf-8') as f:
                json.dump(serialized, f, ensure_ascii=False)
        break

