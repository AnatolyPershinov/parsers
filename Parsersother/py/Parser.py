#!/usr/bin/env python
# coding: utf-8

# In[1]:


import openpyxl
import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import pandas as pd
import numpy as np

import os
import json
import re
import traceback

import jpype
import asposecells
jpype.startJVM()    
from asposecells.api import *


# In[2]:


# import warnings
# warnings.filterwarnings("ignore", category=UserWarning)


# In[3]:


def get_location_by_name(sheet: openpyxl.worksheet.worksheet.Worksheet, target_str: list, 
                         min_row = 1, max_row = None, min_col = 1, max_col = None):
    """Function for finding the position on a page of a cell containing a substring from a list
    Argi:
        sheet (openpyxl.worksheet.worksheet.Worksheet): target sheet in excel file
        target (list): list of substrings
    Returns:
        list: [position_row, position_column, coordinate_of_cell]
    """
    if not max_row:
        max_row = sheet.max_row
    if not max_col:
        max_col = sheet.max_column
        
    target_pos =  []
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            value = sheet.cell(row=row, column = col).value
            coordinate = sheet.cell(row=row, column = col).coordinate
            if value: 
                for name in [var.lower() for var in target_str]:
                    if name == str(value).lower():
                        target_pos.append([row-1, col, coordinate])
    return target_pos


# In[4]:


def get_width(sheet, coordinate: str):
    """
    Function to find a maximal level of multiindex header across rows
    """
    for MergedCell in sheet.merged_cells.ranges:
        if coordinate in MergedCell:
            x1, x2 = str(MergedCell).split(":")

            return sheet[x2].row - sheet[x1].row + 1


# In[5]:


def get_length(sheet, coordinate: str):
    """
    Function to find a maximal level of multiindex header across columns
    """
    for MergedCell in sheet.merged_cells.ranges:
        if coordinate in MergedCell:
            x1, x2 = str(MergedCell).split(":")

            return sheet[x2].column - sheet[x1].column + 1


# In[6]:


def get_start_point(sheet, initial_coords, excluding:list = [], window:int=10):
    """
    Function to get a coordinate of first not-none row in defining column
    Args:
    sheet
    initial_coords: int, point to start search with
    window: int, range of searching
    structure_shift: int, structure shift due to change of structure (or heurstic)
    excluding: List[str], list of excluding markers
    
    """
    finish = initial_coords[0] + window
    column = sheet[get_column_letter(initial_coords[1])]

    for index, cell in enumerate(column):
        if initial_coords[0] < index <= finish:
            if cell.value and all(f not in str(cell.value) for f in excluding):
                return [cell.row, cell.column, cell.value]
    return None


# In[27]:


def get_sheets(path: str, markers: list):
    """
    Function to get workbooks from sheets
    Args:
    Markers: List[str], engrams searched in names of sheets
    
    """
    if path.endswith(".xlsb"):
        print("XLSB file")
        workbook = Workbook(path)
        path = path.replace(".xlsb", ".xlsx")
        workbook.save(path)
    wb = load_workbook(path, data_only=True, allow_none=True)
    wb.active
    sheetnames = wb.sheetnames
    
    workbooks = []
    for marker in markers:
        if any(marker in sheetname for sheetname in sheetnames):        
            sheet = wb[[sheet for sheet in sheetnames if marker.lower() in sheet.lower()][0]]
            workbooks.append(sheet)
        else:
            print(False)
    return workbooks


# In[8]:


def get_header(sheet, start, width, fmt="%Y-%m-%d", bounds:list = []):
    """
    start: int, index of row
    width: int, max level of multiindex header
    fmt: str, date format
    bound: List[datetime.datetime], time bounds for header
    """
    calendar = {m:n for m,n in zip(["январь", "февраль", "март", "апрель", "май", "июнь", "июль", 
                                    "август", "сентябрь", "октябрь", "ноябрь", "декабрь"], range(1, 13))}
    
    merged = sheet.merged_cells.ranges
    header = []
    
    for cell in sheet[start+1]:
#         print(cell.value)
        if cell.value:
            if get_width(sheet, cell.coordinate) == width:
                if isinstance(cell.value, datetime.datetime): # datetime without range
                    header.append(cell.value.strftime(fmt))
                else:
                    header.append(cell.value)
            else:
                if not isinstance(sheet[cell.coordinate].value, datetime.datetime) and get_length(sheet, cell.coordinate):
                    l = get_length(sheet, cell.coordinate)
                    if cell.value.lower() in calendar.keys(): # case with date in str format
                        m = calendar.get(cell.value.lower())
                        for i in range(l):
                            d = int(sheet[get_column_letter(cell.column + i)][cell.row].value)
                            try:
                                i_date = datetime.datetime(2015, m, d)
                            except ValueError:
                                print("Несуществующий день:", m, d)
                                continue
                            if not i_date <= bounds[1]:
                                break
                            else:
                                print(i_date)
                                header.append(i_date.strftime(fmt))
                    else:
                        for i in range(1, l):
                            header.append(" ".join([cell.value, sheet[get_column_letter(cell.column)][cell.row].value]))
                            header.append(" ".join([cell.value, sheet[get_column_letter(cell.column + l - 1)][cell.row].value]))
                elif get_length(sheet, cell.coordinate): # datetime
                    l = get_length(sheet, cell.coordinate)
                    for i in range(l):
                        if not sheet[get_column_letter(cell.column + i)][cell.row].value: # if date range greater than days
                            header.append("unknown" + str(i))
                        elif not sheet[get_column_letter(cell.column + i)][cell.row].value <= bounds[1]:
                            break
                        else:
                            header.append(
                                sheet[get_column_letter(cell.column + i)][cell.row].value.strftime(fmt)
                            )
                else:
                    header.append(cell.value)
        else: # is it None or column without any name?
            if get_start_point(sheet, initial_coords=[cell.row, cell.column], window=500) or cell.column == 1:
                if isinstance(get_start_point(sheet, initial_coords=[cell.row-1, cell.column], window=500)[2], str):
                    continue
                else:
                    header.append('unknown' + str(cell.column))
            
    return [i.replace("\n", "") for i in header if type(i) == str]


# In[9]:


def parse_rows(sheet, start: list, header: list):
    rows = [row for row in sheet.rows][start[0]-1:]
    final = {}
    for row in rows:
        for key, cell in zip(header, row):
            value = cell.value

            if not key in final.keys():
                final[key] = []

            final[key].append(value)
    return final


# In[10]:


def padding(l: list, end: int, fill=np.nan):
    return (l + end * [fill])[:end]

def msg_to_df(sheet_msg, bounds):
    defining_name = 'наименование работ' # the most informative cell
    initial_coordinates = get_location_by_name(sheet_msg, [defining_name])[0]
    
    w = get_width(sheet_msg, initial_coordinates[2])
    
    start = get_start_point(sheet_msg, initial_coordinates)
    
    header = get_header(sheet_msg, start = initial_coordinates[0],
                        width=w, bounds = bounds)
    final = parse_rows(sheet_msg, start, header)
    
#     for k,v in final.items():
#         print(k, len(v))
    
#     sizes = [len(a) for a in final.values()]
#     if not sizes.count(sizes[0]) == len(sizes): # case with wrong nans in final.values()
#         maximum = max(len(a) for a in final.values())
#         for key, column in final.items():
#             final[key] = padding(column, maximum)

    sizes = [len(a) for a in final.values()]
    if not sizes.count(sizes[0]) == len(sizes): # case with wrong nans in final.values()
        minimum = min(len(a) for a in final.values())
        for key, column in final.items():
            final[key] = column[0:minimum]
        
    msg = pd.DataFrame(final)
#     msg = msg.dropna(axis=1, how='all')
    msg = msg.drop(["план/факт"], axis=1)    
    return msg


# In[11]:


def res_to_df(sheet_res, bounds):
    defining_name_resourses = 'ресурсы'
    ic = get_location_by_name(sheet_res, [defining_name_resourses])[0]
    
    header = get_header(sheet_res, start = ic[0], width=get_width(sheet_res, ic[2]), bounds=bounds)
    
    start = get_start_point(sheet_res, ic)
    
    rows = [row for row in sheet_res.rows][start[0]-1:]
    final = {}
    for row in rows:
        for key, cell in zip(header, row):
            value = cell.value

            if not key in final.keys():
                final[key] = []

            final[key].append(value)
    
    sizes = [len(a) for a in final.values()]
    if not sizes.count(sizes[0]) == len(sizes):
        maximum = max(len(a) for a in final.values())
        for key, column in final.items():
            final[key] = padding(column, maximum)
            
    res = pd.DataFrame(final)
    res = res.dropna(axis=1, how='all')
    
    return res   


# In[12]:


def check_row_onNan(row):
    ONCHECK = ["Всегопо проекту", "С начала строительства  план", "С начала строительства  факт"]
    if all(pd.isna(a) for a in row[ONCHECK]):
        return True
    else:
        return False


# In[13]:


def to_date(row):
    try: 
        row = str(row)
    except:
        print("Unserializable row", row)
        return None
    
    if isinstance(row, datetime.datetime):
        try: # NaT
            o = row.strftime("%Y.%m.%d")
            return o
        except:
            return str(row)  
    
    calendar = {m:n for m,n in zip(["январь", "февраль", "март", "апрель", "май", "июнь", "июль", 
                                    "август", "сентябрь", "октябрь", "ноябрь", "декабрь"], range(1, 13))}
    pattern = "(\w{2,7}).?\s?(\d\d)"
    if re.findall(pattern, str(row)):
        monat, year = re.findall(pattern, row)[0]
    else:
        return None
    for month, m in calendar.items():
        if monat in month:
            return datetime.datetime(2000+int(year), m, 1).strftime("%Y.%m.%d")


# In[14]:


def validate_numeric(data):
    if isinstance(data, str):
        if "%" in data:
            return float(data.replace("%", ""))
        else:
            return np.nan
    else:
        return data


# In[15]:


# for i in msg.columns:
#     print(i)


# In[16]:


def msg_to_json(msg):
    operation_package_start = False
    init_uw = True
    upper_works_start = True
    level = -1
    
    arbeit_index = 0
    
    
    upper_works = []
    parsed = {"work": []}

    for _, mrow in msg.iterrows():
        if not mrow['Наименование работ']:
            if all(pd.isna(a) for a in mrow):
                if operation_package_start:
                    operation_package_start = False
                    level = -1
                else:
                    continue

            else: # fact
                if not operation_package_start:
                    continue

                else: # Fact dates
                    days_data = mrow.dropna()
                    for date, value in zip(days_data.index, days_data):
                        if re.search("\d\d-\d\d-\d\d", date):
                            for index in range(len(package['work_data']['progress'])):
                                if date in package['work_data']['progress'][index].keys():
                                    package['work_data']['progress'][index][date]['fact'] = value
                    operation_package_start = False
                    level = -1

        else: # info, upper work
            if check_row_onNan(mrow):
                if init_uw:
                    upper_works.append(mrow["Наименование работ"])
#                     print('Up 1 -> ', upper_works)
                    continue
                elif not operation_package_start: # after package start
                    upper_works_start = True
                    try:
                        upper_works[level] = mrow["Наименование работ"]
                    except IndexError:
                        break # in case of huge gap in defining column break the loop
                    if not level == -1: # level shift
                        upper_works.pop(level) 
                        upper_works.append(mrow["Наименование работ"])
        
                    level -= 1                      
            else:          
                upper_works_start = False
                package = {}
                operation_package_start = True
                init_uw = False

#                 print('Up 2 -> ', upper_works)

                ff = upper_works[:]
                package["upper works"] = ff

                package["work title"] = mrow["Наименование работ"]
                package["work id"] = arbeit_index
                package["measurements"] = mrow["Ед.изм"]
                package["amount"] = mrow['Всегопо проекту']
                package["work_data"] = {
                    "start_date": {
                        "plan": to_date(mrow["Начало План"]) if mrow["Начало План"] else None,
                        "estimate": None,
                        "fact": to_date(mrow["Начало Факт"]) if mrow["Начало Факт"] else None
                        },
                    "stop_date": {
                          "plan": to_date(mrow["Окончание План"]) if mrow["Окончание План"] else None,
                          "estimate": None,
                          "fact": to_date(mrow["Окончание Факт"]) if mrow["Окончание Факт"] else None
                        },
                    "complite_state_value": {
                          "plan": validate_numeric(mrow['С начала строительства  план']),
                          "fact": validate_numeric(mrow['С начала строительства  факт'])
                        },
                    "complite_state_perc": {
                          "plan": validate_numeric(mrow["% вып. с начала строительства От плана"]),
                          "fact": validate_numeric(mrow["% вып. с начала строительства От всего объема"])
                        },
                    "current_remain_perc": None,
                    "current_remain_value": None,
                    "whole_remain_perc": None,
                    "whole_remain_value": None,
                    
                    "current_remain": None,
                    "whole_remain": None,
                    "mounth_complite_value": {
                        "plan": validate_numeric(mrow["С начала месяца план"]),
                        "fact": validate_numeric(mrow["С начала месяца факт"])},
                    "mounth_complite_perc": {
                        "plan": None,
                        "fact": validate_numeric(mrow["% вып. плана с начала месяца"])
                        },
                    "progress": []
                }
                for date, value in zip(mrow.index, mrow):
                    if re.search("\d\d-\d\d-\d\d", date):
                        package["work_data"]["progress"].append({date: {"plan": value, "fact": None}})
                parsed['work'].append(package)
                arbeit_index += 1
    return parsed


# In[17]:


def res_to_json(res):
    res_name = "human"
    idx = 0
    
    package_start = False
    
    parsed = []
    for _, rrow in res.iterrows(): 
        if not pd.isna(rrow['Ресурсы']):
            if rrow['Ресурсы'] == "Итого":
                continue
            
            if "техника" in rrow['Ресурсы'].lower():
                res_name = "equipment"
                continue
            if "персонал" in rrow['Ресурсы'].lower():
                res_name = "human"
                continue
            if 'ресурсы' in rrow['Ресурсы'].lower():
                res_name = "human" if rrow['Ресурсы'].split()[0].lower() == "людские" else "equipment"
#                 res_name = None
                continue
            
            if "Составил" in rrow['Ресурсы']:
                continue
            
            package_start = True
            rpack = {
              "resource_id": idx,
              "resource_name": rrow['Ресурсы'],
              "type": res_name,
              "progress": []
                }
            parsed.append(rpack) 
            idx+=1
            days_data = rrow.dropna()
            for date, value in zip(rrow.index, rrow):
                if re.search("\d\d-\d\d-\d\d", date):
                    rpack["progress"].append({date: {"plan": value, "fact": None}})
        else:
            if not package_start:
                continue
            for date, value in zip(rrow.index, rrow):
                if re.search("\d\d-\d\d-\d\d", date):
                    for index in range(len(rpack['progress'])):
                        if date in rpack['progress'][index].keys():
                            rpack['progress'][index][date]['fact'] = value
            package_start = False
    return parsed


# In[18]:


def get_months_range(path):
    calendar = {m:n for m,n in zip(["январь", "февраль", "март", "апрель", "май", "июнь", "июль", 
                                    "август", "сентябрь", "октябрь", "ноябрь", "декабрь"], range(1, 13))}
    
    info = path.split('\\')[-2]
    month = info.split()[1]
    

    m = calendar[month.lower()]
    
    if m > 12:
        m -= 11
    
    last = datetime.date(2020, m+1 if m < 12 else 1, 1) - datetime.timedelta(days=1)
    last_day = last.day
    
    # range may vary
    bounds = [0, datetime.datetime(2020, m, last_day)]
    
    return bounds


# In[19]:


def find_convert_dt(s):
    pattern = "(\d\d).(\d\d).*(\d{2,4})"
#     candidates = re.findall(pattern, s)[0]
#     print(candidates)
    d, m, y = map(int, re.findall(pattern, s)[0])
    if y < 2000:
        return datetime.datetime(2000+y, m, d).strftime("%d.%m.%Y")
    else:
        return datetime.datetime(y, m, d).strftime("%d.%m.%Y")


# In[20]:


# find_convert_dt("R13- МСГ УКПГ КХМ_31.12. 2017.xlsb")


# In[21]:


PATH = r"C:\Users\Roman\Desktop\Project 234\СМГ УКПГ Н Порт\МСГ УКПГ\МСГ УКПГ"

jahre = [a for a in os.listdir(PATH) if a[0].isdigit()]

monate = []
for jahr in jahre:
    monate.append([f"{jahr}\\{a}" for a in os.listdir(os.path.join(PATH, jahr))])
    
files = {}
for i in monate:
    for monat in i:
        files[os.path.relpath(monat)] = [os.path.join(PATH, monat, file) for file in os.listdir(os.path.join(PATH, monat))
                                        if not os.path.isdir(os.path.join(PATH, monat, file))]
for d, jahr in files.items():
    if d.startswith("2021"):
        break
    if d.startswith(("2015", "2016", "2017", "2018", "2019")):
        continue
    z = tuple(a.capitalize() + " МСГ 2019" for a in []) # done
    if d.endswith(z):
        continue
    else:
        for path in jahr:
#             print(path)
            print("\\".join(path.split('\\')[-3:]))
            try:
                sheet_msg, sheet_res = get_sheets(path=path, markers=['МСГ', 'Ресурсы'])
                msg = msg_to_df(sheet_msg, bounds=get_months_range(path))
                
                res = res_to_df(sheet_res, bounds=get_months_range(path))

                serialized = msg_to_json(msg)
                serialized['resource'] = res_to_json(res)
                serialized['file_name'] = "\\".join(path.split('\\')[-3:])
                name = "Н.Порт_МСГ_УГПК_" + find_convert_dt(path.split('\\')[-1])
                print("TimeOp ->", find_convert_dt(path.split('\\')[-1]))
                with open(f"{name}.json", 'w', encoding='utf-8') as f:
                    json.dump(serialized, f, ensure_ascii=False)
                print("-----")
            except Exception as ex:
                print(traceback.format_exc())
                continue
#                 break
#         break


# In[28]:


path = r"C:\Users\Roman\Desktop\Project 234\СМГ УКПГ Н Порт\МСГ УКПГ\МСГ УКПГ\2019 МСГ\03 Март МСГ 2019\R-13 МСГ УКПГ КХМ_21.03.2019.xlsb"
sheet_msg, sheet_res = get_sheets(path=path, markers=['МСГ', 'Ресурсы'])

# msg = msg_to_df(sheet_msg, bounds=get_months_range(path))

# res = res_to_df(sheet_res, bounds=get_months_range(path))

# serialized = msg_to_json(msg)
# serialized['resource'] = res_to_json(res)
# serialized['file_name'] = "\\".join(path.split('\\')[-3:])
# name = "Н.Порт_МСГ_УГПК_" + find_convert_dt(path.split('\\')[-1])
# print("TimeOp ->", find_convert_dt(path.split('\\')[-1]))
# with open(f"{name}.json", 'w', encoding='utf-8') as f:
#     json.dump(serialized, f, ensure_ascii=False)


# In[ ]:


# serialized


# In[29]:


with open(r"Н.Порт_МСГ_УГПК_16.03.2019.json", encoding="utf8") as f:
    a = json.load(f)
a


# In[ ]:




