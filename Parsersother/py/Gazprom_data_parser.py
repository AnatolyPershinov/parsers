#!/usr/bin/env python
# coding: utf-8

# In[1]:


import openpyxl
import pandas as pd
from pyxlsb import open_workbook as open_xlsb
from datetime import datetime
import os
from datetime import timedelta
import xlrd
from pyxlsb import open_workbook
import copy
import json
from tqdm.notebook import tqdm as tq


# In[2]:


months = {'Январь': 1, 'Февраль': 2,'Март':3, 'Апрель':4, 'Май':5, 'Июнь':6, 'Июль':7, 'Август':8, 'Сентябрь':9, 'Октябрь':10, 'Ноябрь':11, 'Декабрь':12}
years = ['2015', '2016', '2017']

estimate_lst = ['Ожид']


# In[ ]:





# In[3]:


def recursive_flatten_generator(array):
    lst = []
    for i in array:
        if isinstance(i, list):
            lst.extend(recursive_flatten_generator(i))
        else:
            lst.append(i)
    return lst

def estimate_in(sheet, row):
    if 'Ожид' in sheet.row_values(row, start_colx=0, end_colx=None):
        return True
    else:
        return False

def format_date(date):
    new_date = xlrd.xldate_as_tuple(date, wb.datemode)
    format_date = str(new_date[2])+'.'+str(new_date[1])+'.'+str(new_date[0])
    return format_date
#format_date(42387)

def get_month_num(sheet, months, filename):
    if any(months.keys()) in sheet.col_values(month_col):
        #print(True)
        month = list(set(months.keys()) & set(sheet.col_values(month_col)))[0]
        month_num = months[month]
    return month_num

def get_year(sheet, title_col):
    s = sheet.cell(title_row, title_col).value
    #print(s.split())
    if '2015' in s.split():
        year = 2015
    elif '2016' in s.split():
        year = 2016
    elif '2017' in s.split():
        year = 2017
    else:
        print('Year mistake')
    return year

def get_day_id(sheet, day, m):
    year = 2016
    return str(day)+'.'+str(m)+'.'+str(year)


# In[ ]:





# ## Словарь индексов и работ

# In[4]:


#Словарь работ из файла

def work_dict(sheet, work_row, id_col, work_name_col):
    keys = []
    values = []
    if 'Наименование этапа строительства' in sheet.col_values(work_name_col):
        k = 6
    else:
        k=4
    for i in range(work_row+k, sheet.nrows):
        keys.append(str(sheet.cell_value(i, id_col)))
        values.append(sheet.cell(i, work_name_col).value)
    if not '№ п/п' in keys:
        keys.append('№ п/п')
    cut = keys.index('№ п/п')
    keys = [item for item in keys[:cut] if item!='']
    keys = [item.replace('.0','') for item in keys]
    values = [item for item in values if item!='' and item!='Наименование этапа строительства']
    values = [item for item in values[:cut] if item!='']
    work_dict = dict(zip(keys, values))
    #print(work_dict)
    return work_dict

def work_dict_dspmk(sheet, work_row, id_col, work_name_col):
    keys = []
    values = []
#     if 'Наименование этапа строительства' in sheet.col_values(work_name_col):
#         k = 6
#     else:
    k=4
    for i in range(work_row+k, sheet.nrows):
        keys.append(sheet.cell_value(i, id_col))
        values.append(sheet.cell_value(i, work_name_col))
    if not '№ п/п' in keys:
        keys.append('№ п/п')
    cut = keys.index('№ п/п')
    keys = [str(item) for item in keys[:cut]]
    keys = [item.replace('.0','') for item in keys]
    values = [str(item) for item in values[:cut]]
    work_dict = dict(zip(keys, values))
    return work_dict

def work_dict_gpnv(sheet, work_row, work_name_col):
    id_col = 0
    keys = []
    values = []

    for i in range(work_row+4, sheet.nrows):
        values.append(sheet.cell(i, work_name_col).value)
        keys.append(sheet.cell_value(i, id_col))
        keys.append(sheet.cell_value(i, id_col+1))
        keys = [item for item in keys if 'gpnv' not in str(item) and item!='']

    cut = keys.index('№ п/п')
   
    keys = [item for item in keys[:cut-1] if item!='']
    keys = [item.replace('.0','') for item in keys]
    values = [item for item in values if item!='' and item!='Наименование этапа строительства']
    values = [item for item in values[:cut-1]]
    work_dict = dict(zip(keys, values))
    #print(work_dict)
    return work_dict

def work_dict_gks(sheet, start_row, stop_row, id_col):
    keys = []
    values = []
    if 'Наименование этапа строительства' in sheet.col_values(work_name_col):
        k = 6
    else:
        k=4
    for i in range(start_row+1, stop_row):
        keys.append(str(sheet.cell_value(i, id_col)))
        values.append(sheet.cell(i, work_name_col).value)
    if not '№ п/п' in keys:
        keys.append('№ п/п')
    #print(values)
    cut = keys.index('№ п/п')
    keys = [item for item in keys[:cut] if item!='']
    keys = [item.replace('.0','') for item in keys]
    values = [item for item in values if item!='' and item!='Наименование этапа строительства']
    values = [item for item in values[:cut] if item!='']
    dict_1 = dict(zip(keys, values))
    dict_1
    return dict_1

def work_dict_17(sheet, work_row, id_col, work_name_col, stop):
    work_dict = dict()
    if 'Наименование этапа строительства' in sheet.col_values(work_name_col):
        k = 5
    else:
        k=3    
    for i in range(work_row+k, stop):
        key = str(sheet.cell_value(i, id_col))
        value = sheet.cell_value(i, work_name_col)
        if key!='' and value!='':
            work_dict.update({'{}'.format(key): value})
    keys = list(work_dict.keys())
    values = work_dict.values()
    if not '№ п/п' in keys:
        keys.append('№ п/п')
    cut = keys.index('№ п/п')
    keys = [item for item in keys[:cut] if item!='']
    keys = [item.replace('.0','') for item in keys]
    values = [item for item in values if item!='' and item!='Наименование этапа строительства']
    values = [item for item in values[:cut] if item!='']
    work_dict = dict(zip(keys, values))

    return work_dict
#Выделение более высоких узлов
#work_id = '4.5.1.3'
def get_upper_works(work_id, work_dict):
    level = work_id.count('.')+1
    id_lst=[]
    for i in range(1,level):
        new_id = work_id.split('.')
        new_id = '.'.join(new_id[:-1])
        id_lst.append(new_id)
        work_id = new_id
    #print(id_lst)
    wrk_lst=[]
    for item in id_lst:
        if item in work_dict.keys():
            try:
                item = int(item)
            except:
                item = item        
            wrk = work_dict[str(item)]
            wrk_lst.append(str(item)+' '+str(wrk))
    return wrk_lst

def get_upper_works_gks(work_id, work_name, dicts):
    work_dict = dicts[0]
    for item in dicts:
        if work_name in item.values():
            work_dict = item
    level = work_id.count('.')+1
    id_lst=[]
    for i in range(1,level):
        new_id = work_id.split('.')
        new_id = '.'.join(new_id[:-1])
        id_lst.append(new_id)
        work_id = new_id
    #print(id_lst)
    wrk_lst=[]
    for item in id_lst:
        if item in work_dict.keys():
            try:
                item = int(item)
            except:
                item = item        
            wrk = work_dict[str(item)]
            wrk_lst.append(str(item)+' '+str(wrk))
    return wrk_lst

def get_rows(sheet, work_name_col, stop_row):
    rows=[]
    try:
        for row in range(work_row+4, stop_row):
            wrk = sheet.cell(row, work_name_col).value
            unit = sheet.cell(row, unit_col).value
            if wrk!='' and unit!='':
                rows.append(row)
    except:
        try:
            for row in range(work_row+5, stop_row):
                wrk = sheet.cell(row, work_name_col).value
                unit = sheet.cell(row, unit_col).value
                if wrk!='' and unit!='':
                    rows.append(row)
        except:
            for row in range(work_row+6, stop_row):
                wrk = sheet.cell(row, work_name_col).value
                unit = sheet.cell(row, unit_col).value
                if wrk!='' and unit!='':
                    rows.append(row)
            
    return rows


# In[ ]:





# In[ ]:





# In[ ]:





# In[5]:


#выводит колонку по номеру
#for row_idx in range(0, sheet.nrows):
#    print(row_idx, str(sheet.cell(row_idx, 49).value))
    
#выводит ряд по номеру
#for col_idx in range(0, sheet.ncols):
#    print(col_idx, str(sheet.cell(10, col_idx).value))

#sheet.cell(fact_row, first_day+14).value


# In[ ]:





# In[ ]:





# ## Парсинг прогресса

# In[5]:


def make_progress(sheet, ROW, days_row, days_col, months, month_col, title_col, comment_col, work_name_col):
    work = sheet.cell(ROW, work_name_col).value
    progress = []
    if work!='':
        day0 = 1
        #print(day0)
        N = check_days(m)
        for n in range(0,N):
            day = day0+n
            day_id = get_day_id(sheet, day, m)
            try:
                plan = sheet.cell(ROW, days_col+1+n).value
            except:
                plan = 0
            if not isinstance(plan, float) or not isinstance(plan, int):
                plan = 0
    
            try:
                fact = sheet.cell(ROW+1, days_col+1+n).value
                if fact == '':
                    fact = 0
            except:
                fact = 0
            if not isinstance(fact, float) or not isinstance(fact, int):
                fact = 0

            prog = {str(day_id):{'plan':plan, 'fact': fact}}
            progress.append(prog)
    return progress


# In[ ]:





# ## Поиск колонок

# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:





# In[6]:


def gpnv_in(sheet):
    if 'gpnv' in ' '.join(map(str,sheet.col_values(0))) or 'gpnv' in ' '.join(map(str,sheet.col_values(1))):
        return True
    else:
        return False    
    
def dspmk_in(file):
    if ' ДСПМК ' in file or ' СЭМ ' in file:
        dspmk = True
    else:
        dspmk = False
    return dspmk


# In[7]:


def get_info(sheet, ROW, dicts=False, dspmk=False, **params):
    if gpnv_in(sheet):
        wrk_dict = work_dict_gpnv(sheet, params['work_row'], params['work_name_col'])
    elif dspmk:
        wrk_dict = work_dict_dspmk(sheet, params['work_row'], params['id_col'], params['work_name_col'])
    else:    
        wrk_dict = work_dict_17(sheet, params['work_row'], params['id_col'], params['work_name_col'], params['stop'])
    
    for col_idx in range(0, sheet.ncols):
        work_title = str(sheet.cell(ROW, params['work_name_col']).value)
            
        work_id = str(sheet.cell(ROW, params['id_col']).value)
        if work_id =='':
            work_id = None
            
        #work_id = sheet.cell(ROW, params['id_col']).value
        
        if dicts:
            upper_works = get_upper_works_gks(work_id, work_title, dicts)
        else:
        if work_id:
            upper_works = get_upper_works(work_id, wrk_dict)
        else:
            upper_works = None
    
        unit = str(sheet.cell(ROW, params['unit_col']).value)
            
        if isinstance(sheet.cell_value(ROW, params['volume_col']), str):
            amount = float(sheet.cell_value(ROW, params['volume_col']).replace(',','.'))
        else:
            amount = sheet.cell_value(ROW, params['volume_col'])
        
        try:    
            start_plan = format_date(sheet.cell(ROW, params['start_col']).value)
        except:
            start_plan = None
        if estimate_in(sheet,params['start_row']+1):
            try:
                start_est = format_date(sheet.cell(ROW, params['start_col']+1).value)
            except:
                start_est = None
        else:
            start_est = None
        try:
            start_fact = format_date(sheet.cell(ROW, params['start_col']+2).value)
        except:
            start_fact = None
        
        start_date = {"plan": start_plan,
                      "estimate": start_est,
                      "fact": start_fact}
        try:
            stop_plan = format_date(sheet.cell(ROW, params['stop_col']).value)
        except:
            stop_plan = None
        if estimate_in(sheet,params['stop_row']+1):
            try:
                stop_est = format_date(sheet.cell(ROW, params['stop_col']+1).value)
            except:
                stop_est = None
        else:
            stop_est = None
        try:
            stop_fact = format_date(sheet.cell(ROW, params['stop_col']+2).value)
        except:
            stop_fact = None
        stop_date = {
                      "plan": stop_plan,
                      "estimate": stop_est,
                      "fact": stop_fact}
    
        if not isinstance(sheet.cell_value(ROW, params['complete_col']), str):
            complite_state_plan_perc = sheet.cell_value(ROW, params['complete_col'])/amount*100
        else:
            complite_state_plan_perc = None
            
        complite_state_fact_perc = sheet.cell_value(ROW, params['complete_col']+1)/amount*100 
        complite_state_perc = {
                      "plan": complite_state_plan_perc,
                      "fact": complite_state_fact_perc}
        
        if not isinstance(sheet.cell_value(ROW, params['complete_col']), str):
            complite_state_plan_value = sheet.cell_value(ROW, params['complete_col'])
        else:
            complite_state_plan_value = None
        
        complite_state_fact_value = sheet.cell_value(ROW, params['complete_col']+1)
        complite_state_value = {
                      "plan": complite_state_plan_value,
                      "fact": complite_state_fact_value}
    
        mounth_complite_plan_perc = sheet.cell_value(ROW, params['mounth_col'])/amount*100
        mounth_complite_fact_perc = sheet.cell_value(ROW, params['mounth_col']+1)/amount*100
        mounth_complite_perc = {
                      "plan": mounth_complite_plan_perc,
                      "fact": mounth_complite_fact_perc}
        
        mounth_complite_plan_value = sheet.cell_value(ROW, params['mounth_col'])
        mounth_complite_fact_value = sheet.cell_value(ROW, params['mounth_col']+1)
        mounth_complite_value = {
                      "plan": mounth_complite_plan_value,
                      "fact": mounth_complite_fact_value}
    
        try:
            current_remain_perc = 100 - sheet.cell_value(ROW, params['current_col'])
        except:
            current_remain_perc = 0
        current_remain_value = 0
        if not isinstance(current_remain_perc, float) or not isinstance(current_remain_perc, int):
            current_remain_perc = 0
            
             
        whole_remain_perc = 100 - sheet.cell_value(ROW, params['whole_col'])
        whole_remain_value = sheet.cell(ROW, params['whole_col']-1).value
        
        if not isinstance(whole_remain_value, float) or not isinstance(whole_remain_value, int):
            whole_remain_value = complite_state_plan_value
            
        if not isinstance(whole_remain_perc, float) or not isinstance(whole_remain_perc, int):
            whole_remain_perc = 0
        
        comment = sheet.cell(ROW, params['comment_col']).value
           
    progress = make_progress(sheet, ROW, params['days_row'], params['days_col'], params['months'], 
                             params['month_col'], params['title_col'], params['comment_col'], params['work_name_col'])
        
    work_data = {"start_date": start_date,
                 "stop_date": stop_date,
                 "complite_state_perc": complite_state_perc,
                 'complite_state_value': complite_state_value,
                 "current_remain_perc": current_remain_perc,
                 'current_remain_value': current_remain_value,
                 "whole_remain_perc": whole_remain_perc,
                 "whole_remain_value": whole_remain_value,
                 "mounth_complite_perc": mounth_complite_perc,
                 "mounth_complite_value": mounth_complite_value,
                 'comments': comment,
                 "progress": progress
                }
    d = {'work_title': work_title,
         'work_id': work_id,
         "upper works": upper_works,
         'measurements': unit,
         "amount": amount,
         "work_data": work_data
        }
    return d
    


# ## Ресурсы

# In[31]:


def type_(ROW):
    type_ = None
    if 'Наименование и марка техники (механизма), оборудования' in sheet.row_values(ROW):
        type_ = 'equipment'
    elif 'Наименование должностей, профессий' in sheet.row_values(ROW):
        type_ = 'human'
    return type_


# In[32]:


def check_days(m):
    if m==1 or m==3 or m==5 or m==7 or m==8 or m==10 or m==12:
        n = 31
    elif m==2:
        n = 29
    elif m==4 or m==6 or m==9 or m==11:
        n = 30
    return n


# In[9]:


def make_progress_res(sheet, ROW, resource_name_row, resource_name_col, days_col,months, month_col, title_col):
    #resource_name_col = 1
    res = sheet.cell(ROW, resource_name_col).value
    progress = []
    if res!='':
        day0 = 1
        #day0 = int(sheet.cell_value(resource_name_row+1, days_col+1))
        #print(day0)
        N = check_days(m)
        for n in range(0,N):
            day = day0+n
            day_id = get_day_id(sheet, day, m)
            try:
                plan = sheet.cell(ROW, days_col+1+n).value
            except:
                plan = 0
            if not isinstance(plan, float) or not isinstance(plan, int):
                plan = 0
            try:
                fact = sheet.cell_value(ROW+1, days_col+1+n)
            except:
                fact = 0
            if not isinstance(fact, float) or not isinstance(fact, int):
                fact = 0
    
            prog = {str(day_id):{'plan':plan, 'fact': fact}}
            progress.append(prog)
    return progress

def get_info_res(sheet, ROW, resource_row, resource_col, **params):
    
    days_col = params['days_col']
    months = params['months']
    month_col = params['month_col']
    title_col = params['title_col']
    
    resource_id = sheet.row_values(ROW)[0]
    resource_name = sheet.row_values(ROW)[resource_col]
    resource_type = type_(resource_row)
    comments = sheet.cell_value(ROW, params['comment_col'])
    
    progress = make_progress_res(sheet, ROW, resource_row, resource_col, days_col, months, month_col, title_col)
   
    resource = {'resource_id': resource_id,
                'resource_name': resource_name,
                'resource_type': resource_type,
                'comments':comments,
                'progress': progress
                }
    return resource


# In[ ]:





# In[ ]:





# In[ ]:





# ## Парсинг файла

# In[ ]:





# In[13]:





# In[16]:





# In[17]:


m = 1
root_path = 'Z://GPN_KIP//parsed//ТИП//2016//2016_1//'
#p = 'Z://GPN_KIP//GPN_KIP//Томский интегрированный проект//Ежесуточная//Ежесуточная//Недельно (суточный)-месячный план-график//2016//2016_04//'

def choose_sheet(wb):
    m = [item for item in wb.sheet_names() if 'январь' in item or 'Январь' in item or '1.16' in item]
    if len(m)>1:
        m = [item for item in m if '2016' in item or '16' in item]
        m = m[0]
    elif len(m)==1:
        m = m[0]
    else:
        print('No sheet')
    return m
p = 'Z://GPN_KIP//GPN_KIP//Томский интегрированный проект//Ежесуточная//Ежесуточная//Недельно (суточный)-месячный план-график//2016//2016_01//'
# paths = os.listdir(p)
# paths = [p+item+'//' for item in paths]
# paths = [item for item in paths if not '.rar' in item]
# len(paths)

paths = os.listdir(p)
file_paths = [p+item for item in paths if '.xls' in item or '.XLS' in item]
folder_paths = [p+item+'//' for item in paths if '.xls' not in item and '.rar' not in item and '.db' not in item and 'wrong' not in item]

new_paths=[]
for item in folder_paths:
    files_list = os.listdir(item)
    f_paths = [item+l for l in files_list]
    new_paths.append(f_paths)
new_paths = recursive_flatten_generator(new_paths)


paths = file_paths + new_paths
paths = [path for path in paths if ' ГКС ' in path or '-ГКС-' in path] #ГКС пока не парсятся
paths = [path for path in paths if not 'Сводка' in path and 'План' not in path]
len(paths)


# In[18]:


paths[:1]


# In[20]:


wrong=[]


# In[25]:


len(paths)


# In[26]:


from tqdm.notebook import tqdm as tq


for path in tq(paths):

    ind = path.rfind('//')
    name = path[ind+2:]
    print(name)
    #files = [file for file in os.listdir(path)]

    #non_files = [file for file in files if '.xls' in file or '.xlsx' in file or '.XLSX' in files]
   

    #print(len(paths))
    something_wrong=[]


    dspmk = dspmk_in(path)
            
    try:
#         filename, file_extension = os.path.splitext(file)
#         print(filename)
        wb = xlrd.open_workbook(path)
        sheet = choose_sheet(wb) 
        #print(sheet)
        sheet = wb.sheet_by_name(sheet)


        stop_rows=[]
        equipment = []
        human = []
        for rowidx in range(sheet.nrows):
            row = sheet.row(rowidx)
            id_col = 0
            for colidx, cell in enumerate(row):
                if cell.value == "Наименование работ" :
                    work_row = rowidx
                    work_name_col = colidx
                    plan_fact_row = work_row+2
                    #print(sheet.cell(work_row, work_name_col).value)
                if cell.value == "Ед. изм." :
                    unit_row = rowidx
                    unit_col = colidx
                    #print(sheet.cell(unit_row, unit_col).value)
                if cell.value == "Кол-во всего по проекту " or cell.value == "Кол-во всего по проекту":
                    volume_row = rowidx
                    volume_col = colidx
                    #print(sheet.cell(volume_row, volume_col).value)
                if cell.value == "Начало" :
                    start_row = rowidx
                    start_col = colidx
                    #print(sheet.cell(start_row, start_col).value)
                if cell.value == "Окончание" :
                    stop_row = rowidx
                    stop_col = colidx
                    #print(sheet.cell(stop_row, stop_col).value)
                if cell.value == "Дни мес.":
                    days_row = rowidx
                    days_col = colidx
                    month_col = days_col+1
                    #print(sheet.cell(days_row, days_col).value)
                if cell.value == "Примечание":
                    comment_row = rowidx
                    comment_col = colidx
                    #print(sheet.cell(comment_row, comment_col).value)
                if cell.value == "общий % выполнения":
                    whole_row = rowidx
                    whole_col = colidx
                    #print(sheet.cell(whole_row, whole_col).value)
                if cell.value == "% выполнения за месяц":
                    current_row = rowidx
                    current_col = colidx
                    #print(sheet.cell(current_row, current_col).value)
                if 'Недельно (суточный)' in str(sheet.cell(rowidx, colidx).value):
                    title_row = rowidx
                    title_col = colidx
                    #print(sheet.cell(title_row, title_col).value)
                if cell.value == "Выполнено с начала строительства":
                    complete_row = rowidx
                    complete_col = colidx
                    #print(sheet.cell(complete_row, complete_col).value)
                if cell.value == "Задание на месяц":
                    mounth_row = rowidx
                    mounth_col = colidx
                    #print(sheet.cell(mounth_row, mounth_col).value)
                if cell.value == 'Всего работ':
                    stop = rowidx
                    stop_rows.append(stop)
                elif cell.value == '№ п/п':
                    stop = rowidx
                    stop_rows.append(stop)
                if cell.value == "Всего работ":
                    row = rowidx
                    col = colidx
                    equipment.append((row, col))
                if cell.value == "Наименование должностей, профессий":
                    row = rowidx
                    col = colidx
                    human.append((row, col))

        if stop_rows:
            stop = stop_rows[0]
        else:
            stop = sheet.nrows

        params = {'work_row':work_row,
                  'work_name_col':work_name_col,
                  'unit_col':unit_col,
                  'volume_col':volume_col,
                  'start_col':start_col,
                  'start_row':start_row, 
                  'stop_row':stop_row,
                  'stop_col': stop_col,
                  'complete_col':complete_col,
                  'month_col':month_col,
                  'current_col':current_col,
                  'whole_col':whole_col,
                  'days_row':days_row,
                  'days_col':days_col,
                  'months':months,
                  'mounth_col':mounth_col,
                  'title_col':title_col,
                  'comment_col':comment_col,
                  'stop': stop,
                  'id_col': id_col
                 }

        params_res = {     
                  'month_col':month_col,
                  'days_col':days_col,
                  'months':months,
                  'mounth_col':mounth_col,
                  'title_col':title_col,
                  'comment_col':comment_col,
                 }   

        f = {"file_name": name}
        #print(f)

        #Работы
        work_lst = []
        ROWS = get_rows(sheet, work_name_col, stop)
        for ROW in ROWS: 
    #         print(ROW)
            wrk = get_info(sheet, ROW, dspmk, **params)
            work_lst.append(wrk)
        #print(work_lst)

        #Ресурсы
        resources = equipment+human
        resources_rows = [item[0] for item in sorted(resources)]
        resources_cols = [item[1] for item in sorted(resources)]
        resources_cols
        stops=resources_rows[1:]
        stops.append(sheet.nrows)

        resource_lst=[]
        ROWS = []
        for i, resource_row in enumerate(resources_rows):
            R = range(resource_row, stops[i])
            ROWS.append(R)

        for i, interval in enumerate(ROWS):
            #print(resources_cols[i])
            for ROW in interval:
                r = get_info_res(sheet, ROW, resources_rows[i], resources_cols[i], **params)          
                resource_lst.append(r)
        #print(resource_lst)

        r_lst = [item for item in resource_lst if item['resource_name']!='Наименование и марка техники (механизма), оборудования']
        r_lst = [item for item in r_lst if item['resource_name']!='Наименование должностей, профессий']
        r_lst = [item for item in r_lst if item['resource_name']!='']
        r_lst = [item for item in r_lst if isinstance(item['resource_name'], str)]


        f.update({'work': work_lst})
        f.update({'resource': r_lst})
        path_out = root_path + name + '.json'
        with open(path_out, "w", encoding="utf-8") as file:
            json.dump(f, file)
        print(f)
    except:
        something_wrong.append(path)
        pass
    wrong.append(something_wrong)


# In[27]:


wrong = recursive_flatten_generator(wrong)
wrong


# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:





# In[13]:





# # Парсинг 2017

# In[84]:


# m = 7
# def choose_sheet(wb):
#     m = [item for item in wb.sheet_names() if 'июль' in item or 'Июль' in item or '07.17' in item]
#     if len(m)>1:
#         m = [item for item in m if '2017' in item or '17' in item]
#         m = m[0]
#     elif len(m)==1:
#         m = m[0]
#     else:
#         print('No sheet')
#     return m


# In[157]:


# p = 'Z://GPN_KIP//GPN_KIP//Томский интегрированный проект//Ежесуточная//Ежесуточная//Недельно (суточный)-месячный план-график//2017//2017//'
# p = p+'2017_07//'
# paths = os.listdir(p)
# file_paths = [p+item for item in paths if '.xls' in item or '.XLS' in item]
# folder_paths = [p+item+'//' for item in paths if '.xls' not in item and '.rar' not in item and '.db' not in item and 'wrong' not in item]

# new_paths=[]
# for item in folder_paths:
#     files_list = os.listdir(item)
#     f_paths = [item+l for l in files_list]
#     new_paths.append(f_paths)
# new_paths = recursive_flatten_generator(new_paths)


# paths = file_paths + new_paths
# # paths = [path for path in paths if ' ГКС ' in path or '-ГКС-' in path] 
# # paths = [path for path in paths if not 'ДСПМК' in path]
# paths = [path for path in paths if not 'Сводка' in path and 'сводка'not in path]
# paths = [item for item in paths if not '.rar' in item]
# len(paths)


# In[ ]:


paths


# In[ ]:


# # path = 'Z://GPN_KIP//GPN_KIP//Томский интегрированный проект//Ежесуточная//Ежесуточная//Недельно (суточный)-месячный план-график//2017//2017//'
# # path = path + '2017_07//смг к отчетам 14.07.17//'
# # filename = 'СМГ Перевод газопровода в реверс 14.07.2017.xlsx'

# for path in tq(paths[:1]):
#     ind = path.rfind('//')
#     filename = path[ind+2:]
#     print(filename)
#     dspmk = dspmk_in(name)
#     wb = xlrd.open_workbook(path)
#     sheet = choose_sheet(wb) 
#     print(sheet)
#     sheet = wb.sheet_by_name(sheet)

#     row_1, row_2, row_3 = 0,0,0
#     obj_rows =[]
#     stop_rows=[]
#     equipment = []
#     human = []
#     for rowidx in range(sheet.nrows):
#         row = sheet.row(rowidx)
#         id_col = 0
#         for colidx, cell in enumerate(row):
#             if cell.value == "Наименование работ" :
#                 work_row = rowidx
#                 work_name_col = colidx
#                 plan_fact_row = work_row+2
#                 print(sheet.cell(work_row, work_name_col).value)
#             if cell.value == "Ед. изм." :
#                 unit_row = rowidx
#                 unit_col = colidx
#                 #print(sheet.cell(unit_row, unit_col).value)
#             if cell.value == "Кол-во всего по проекту " or cell.value == "Кол-во всего по проекту":
#                 volume_row = rowidx
#                 volume_col = colidx
#                 #print(sheet.cell(volume_row, volume_col).value)
#             if cell.value == "Начало" :
#                 start_row = rowidx
#                 start_col = colidx
#                 #print(sheet.cell(start_row, start_col).value)
#             if cell.value == "Окончание" :
#                 stop_row = rowidx
#                 stop_col = colidx
#                 #print(sheet.cell(stop_row, stop_col).value)
#             if cell.value == "Дни мес.":
#                 days_row = rowidx
#                 days_col = colidx
#                 month_col = days_col+1
#                 #print(sheet.cell(days_row, days_col).value)
#             if cell.value == "Примечание":
#                 comment_row = rowidx
#                 comment_col = colidx
#                 #print(sheet.cell(comment_row, comment_col).value)
#             if cell.value == "общий % выполнения":
#                 whole_row = rowidx
#                 whole_col = colidx
#                 #print(sheet.cell(whole_row, whole_col).value)
#             if cell.value == "% выполнения за месяц":
#                 current_row = rowidx
#                 current_col = colidx
#                 #print(sheet.cell(current_row, current_col).value)
#             if 'Недельно (суточный)' in str(sheet.cell(rowidx, colidx).value):
#                 title_row = rowidx
#                 title_col = colidx
#                 #print(sheet.cell(title_row, title_col).value)
#             if cell.value == "Выполнено с начала строительства":
#                 complete_row = rowidx
#                 complete_col = colidx
#                 #print(sheet.cell(complete_row, complete_col).value)
#             if cell.value == "Задание на месяц":
#                 mounth_row = rowidx
#                 mounth_col = colidx
#                 #print(sheet.cell(mounth_row, mounth_col).value)
#             if cell.value == 'Всего работ':
#                 stop = rowidx
#                 stop_rows.append(stop)
#             elif cell.value == '№ п/п':
#                 stop = rowidx
#                 stop_rows.append(stop)
#             if cell.value == "Наименование и марка техники (механизма), оборудования":
#                 row = rowidx
#                 col = colidx
#                 equipment.append((row, col))
#             if cell.value == "Наименование должностей, профессий":
#                 row = rowidx
#                 col = colidx
#                 human.append((row, col))
#             if 'Газокомпрессорная станция' in str(cell.value):
#                 row_1 = rowidx
#                 col_1 = colidx
#                 obj_rows.append(row_1)
#                 #print(sheet.cell(row_1, col_1).value)
#             if 'ВЛ-35 кВ' in str(cell.value):
#                 row_2 = rowidx
#                 col_2 = colidx
#                 #print(sheet.cell(row_2, col_2).value)
#                 obj_rows.append(row_2)
#             if 'Газопровод внешнего транспорта' in str(cell.value):
#                 row_3 = rowidx
#                 col_3 = colidx
#                 obj_rows.append(row_3)
#                 #print(sheet.cell(row_3, col_3).value)


#     work_dict = work_dict_17(sheet, work_row, id_col, work_name_col, stop)


#     #print(dicts)

#     if stop_rows:
#         stop = stop_rows[0]
#     else:
#         stop = sheet.nrows

#     params = {'work_row':work_row,
#                           'work_name_col':work_name_col,
#                           'unit_col':unit_col,
#                           'volume_col':volume_col,
#                           'start_col':start_col,
#                           'start_row':start_row, 
#                           'stop_row':stop_row,
#                           'stop_col': stop_col,
#                           'complete_col':complete_col,
#                           'month_col':month_col,
#                           'current_col':current_col,
#                           'whole_col':whole_col,
#                           'days_row':days_row,
#                           'days_col':days_col,
#                           'months':months,
#                           'mounth_col':mounth_col,
#                           'title_col':title_col,
#                           'comment_col':comment_col,
#                           'stop': stop,
#                           'id_col': id_col
#                         }

#     params_res = {     
#                         'month_col':month_col,
#                         'days_col':days_col,
#                         'months':months,
#                         'mounth_col':mounth_col,
#                         'title_col':title_col,
#                         'comment_col':comment_col,
#                         }   

#     f = {"file_name": filename}
#     #print(f)

#     #Работы
#     work_lst = []
#     ROWS = get_rows(sheet, work_name_col, stop)
#     for ROW in ROWS: 
#         print(ROW)
#         wrk = get_info(sheet, ROW, dspmk, **params)
#         work_lst.append(wrk)
#     #print(work_lst)

#     #Ресурсы
#     resources = equipment+human
#     resources_rows = [item[0] for item in sorted(resources)]
#     resources_cols = [item[1] for item in sorted(resources)]
#     resources_cols
#     stops=resources_rows[1:]
#     stops.append(sheet.nrows)

#     resource_lst=[]
#     ROWS = []
#     for i, resource_row in enumerate(resources_rows):
#         R = range(resource_row, stops[i])
#         ROWS.append(R)

#     for i, interval in enumerate(ROWS):
#         #print(resources_cols[i])
#         for ROW in interval:
#             r = get_info_res(sheet, ROW, resources_rows[i], resources_cols[i], **params)          
#             resource_lst.append(r)
#     #print(resource_lst)

#     r_lst = [item for item in resource_lst if item['resource_name']!='Наименование и марка техники (механизма), оборудования']
#     r_lst = [item for item in r_lst if item['resource_name']!='Наименование должностей, профессий']
#     r_lst = [item for item in r_lst if item['resource_name']!='']
#     r_lst = [item for item in r_lst if isinstance(item['resource_name'], str)]
#     print(r_lst)

#     f.update({'work': work_lst})
#     f.update({'resource': r_lst})
#     path_out = 'Z://GPN_KIP//parsed//ТИП//2017//2017_07//' + name + '.json'
#     with open(path_out, "w", encoding="utf-8") as file:
#         json.dump(f, file)
#     print(f)


# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:


# path = 'Z://GPN_KIP//GPN_KIP//Томский интегрированный проект//Ежесуточная//Ежесуточная//Недельно (суточный)-месячный план-график//2016//'
# path = path + '2016_09//смг к отчетам 19.09.16//'
# file = 'СМГ Ш- ВКС-СПС 19.09.2016г Проверил инженер НТН Череватюк П.П..xls'

# dspmk = dspmk_in(file)
# wb = xlrd.open_workbook(path+file)
# sheet = choose_sheet(wb) 
# print(sheet)
# sheet = wb.sheet_by_name(sheet)
# #print(sheet)

# stop_rows=[]
# equipment = []
# human = []
# for rowidx in range(sheet.nrows):
#     row = sheet.row(rowidx)
#     id_col = 0
#     for colidx, cell in enumerate(row):
#         if cell.value == "Наименование работ" :
#             work_row = rowidx
#             work_name_col = colidx
#             plan_fact_row = work_row+2
#             #print(sheet.cell(work_row, work_name_col).value)
#         if cell.value == "Ед. изм." :
#             unit_row = rowidx
#             unit_col = colidx
#             #print(sheet.cell(unit_row, unit_col).value)
#         if cell.value == "Кол-во всего по проекту " or cell.value == "Кол-во всего по проекту":
#             volume_row = rowidx
#             volume_col = colidx
#             #print(sheet.cell(volume_row, volume_col).value)
#         if cell.value == "Начало" :
#             start_row = rowidx
#             start_col = colidx
#             #print(sheet.cell(start_row, start_col).value)
#         if cell.value == "Окончание" :
#             stop_row = rowidx
#             stop_col = colidx
#             #print(sheet.cell(stop_row, stop_col).value)
#         if cell.value == "Дни мес.":
#             days_row = rowidx
#             days_col = colidx
#             month_col = days_col+1
#             #print(sheet.cell(days_row, days_col).value)
#         if cell.value == "Примечание":
#             comment_row = rowidx
#             comment_col = colidx
#             #print(sheet.cell(comment_row, comment_col).value)
#         if cell.value == "общий % выполнения":
#             whole_row = rowidx
#             whole_col = colidx
#             #print(sheet.cell(whole_row, whole_col).value)
#         if cell.value == "% выполнения за месяц":
#             current_row = rowidx
#             current_col = colidx
#             #print(sheet.cell(current_row, current_col).value)
#         if 'Недельно (суточный)' in str(sheet.cell(rowidx, colidx).value):
#             title_row = rowidx
#             title_col = colidx
#             #print(sheet.cell(title_row, title_col).value)
#         if cell.value == "Выполнено с начала строительства":
#             complete_row = rowidx
#             complete_col = colidx
#             #print(sheet.cell(complete_row, complete_col).value)
#         if cell.value == "Задание на месяц":
#             mounth_row = rowidx
#             mounth_col = colidx
#             #print(sheet.cell(mounth_row, mounth_col).value)
#         if cell.value == "Всего работ":
#             stop = rowidx
#             stop_rows.append(stop)
#         if cell.value == "Наименование и марка техники (механизма), оборудования":
#             row = rowidx
#             col = colidx
#             equipment.append((row, col))
#         if cell.value == "Наименование должностей, профессий":
#             row = rowidx
#             col = colidx
#             human.append((row, col))

# if stop_rows:
#     stop = stop_rows[0]
# else:
#     stop = sheet.nrows

# params = {'work_row':work_row,
#                       'work_name_col':work_name_col,
#                       'unit_col':unit_col,
#                       'volume_col':volume_col,
#                       'start_col':start_col,
#                       'start_row':start_row, 
#                       'stop_row':stop_row,
#                       'stop_col': stop_col,
#                       'complete_col':complete_col,
#                       'month_col':month_col,
#                       'current_col':current_col,
#                       'whole_col':whole_col,
#                       'days_row':days_row,
#                       'days_col':days_col,
#                       'months':months,
#                       'mounth_col':mounth_col,
#                       'title_col':title_col,
#                       'comment_col':comment_col,
#                       'stop': stop,
#                       'id_col': id_col
#                     }

# params_res = {     
#                     'month_col':month_col,
#                     'days_col':days_col,
#                     'months':months,
#                     'mounth_col':mounth_col,
#                     'title_col':title_col,
#                     'comment_col':comment_col,
#                     }   

# f = {"file_name": file}
# #print(f)

# #Работы
# work_lst = []
# ROWS = get_rows(sheet, work_name_col, stop)
# for ROW in ROWS: 
#     print(ROW)
#     wrk = get_info(sheet, ROW, dspmk, **params)
#     work_lst.append(wrk)
# #print(work_lst)

# #Ресурсы
# resources = equipment+human
# resources_rows = [item[0] for item in sorted(resources)]
# resources_cols = [item[1] for item in sorted(resources)]
# resources_cols
# stops=resources_rows[1:]
# stops.append(sheet.nrows)

# resource_lst=[]
# ROWS = []
# for i, resource_row in enumerate(resources_rows):
#     R = range(resource_row, stops[i])
#     ROWS.append(R)

# for i, interval in enumerate(ROWS):
#     #print(resources_cols[i])
#     for ROW in interval:
#         r = get_info_res(sheet, ROW, resources_rows[i], resources_cols[i], **params)          
#         resource_lst.append(r)
# #print(resource_lst)

# r_lst = [item for item in resource_lst if item['resource_name']!='Наименование и марка техники (механизма), оборудования']
# r_lst = [item for item in r_lst if item['resource_name']!='Наименование должностей, профессий']
# r_lst = [item for item in r_lst if item['resource_name']!='']
# r_lst = [item for item in r_lst if isinstance(item['resource_name'], str)]

# f.update({'work': work_lst})
# f.update({'resource': r_lst})
# path_out = 'Z://GPN_KIP//parsed//ТИП//2016//2016_08//' + file + '.json'
# with open(path_out, "w", encoding="utf-8") as file:
#     json.dump(f, file)
# print(f)


# In[ ]:





# In[78]:


m = 9
root_path = 'Z://GPN_KIP//parsed//ТИП//2016//2016_9//'
#p = 'Z://GPN_KIP//GPN_KIP//Томский интегрированный проект//Ежесуточная//Ежесуточная//Недельно (суточный)-месячный план-график//2016//2016_04//'

def choose_sheet(wb):
    m = [item for item in wb.sheet_names() if 'сентябрь' in item or 'Сентябрь' in item or '09.16' in item]
    if len(m)>1:
        m = [item for item in m if '2016' in item or '16' in item]
        m = m[0]
    elif len(m)==1:
        m = m[0]
    else:
        print('No sheet')
    return m
p = 'Z://GPN_KIP//GPN_KIP//Томский интегрированный проект//Ежесуточная//Ежесуточная//Недельно (суточный)-месячный план-график//2016//2016_09//'

paths = os.listdir(p)
file_paths = [p+item for item in paths if '.xls' in item or '.XLS' in item]
folder_paths = [p+item+'//' for item in paths if '.xls' not in item and '.rar' not in item and '.db' not in item and 'wrong' not in item]

new_paths=[]
for item in folder_paths:
    files_list = os.listdir(item)
    f_paths = [item+l for l in files_list]
    new_paths.append(f_paths)
new_paths = recursive_flatten_generator(new_paths)


paths = file_paths + new_paths
paths = [path for path in paths if ' ГКС ' in path or '-ГКС-' in path] #ГКС пока не парсятся
paths = [path for path in paths if not 'Сводка' in path and 'План' not in path]
paths = [path for path in paths if not 'rar' in path  and 'сводка' not in path]
len(paths)


# In[ ]:





# # Парсинг ГКС

# In[79]:


for path in tq(paths):

    ind = path.rfind('//')
    name = path[ind+2:]
    print(name)


    dspmk = dspmk_in(name)
    wb = xlrd.open_workbook(path)
    sheet = choose_sheet(wb) 
    print(sheet)
    sheet = wb.sheet_by_name(sheet)

    row_1, row_2, row_3 = 0,0,0
    obj_rows =[]
    stop_rows=[]
    equipment = []
    human = []
    for rowidx in range(sheet.nrows):
        row = sheet.row(rowidx)
        id_col = 0
        for colidx, cell in enumerate(row):
            if cell.value == "Наименование работ" :
                work_row = rowidx
                work_name_col = colidx
                plan_fact_row = work_row+2
                print(sheet.cell(work_row, work_name_col).value)
            if cell.value == "Ед. изм." :
                unit_row = rowidx
                unit_col = colidx
                #print(sheet.cell(unit_row, unit_col).value)
            if cell.value == "Кол-во всего по проекту " or cell.value == "Кол-во всего по проекту":
                volume_row = rowidx
                volume_col = colidx
                #print(sheet.cell(volume_row, volume_col).value)
            if cell.value == "Начало" :
                start_row = rowidx
                start_col = colidx
                #print(sheet.cell(start_row, start_col).value)
            if cell.value == "Окончание" :
                stop_row = rowidx
                stop_col = colidx
                #print(sheet.cell(stop_row, stop_col).value)
            if cell.value == "Дни мес.":
                days_row = rowidx
                days_col = colidx
                month_col = days_col+1
                #print(sheet.cell(days_row, days_col).value)
            if cell.value == "Примечание":
                comment_row = rowidx
                comment_col = colidx
                #print(sheet.cell(comment_row, comment_col).value)
            if cell.value == "общий % выполнения":
                whole_row = rowidx
                whole_col = colidx
                #print(sheet.cell(whole_row, whole_col).value)
            if cell.value == "% выполнения за месяц":
                current_row = rowidx
                current_col = colidx
                #print(sheet.cell(current_row, current_col).value)
            if 'Недельно (суточный)' in str(sheet.cell(rowidx, colidx).value):
                title_row = rowidx
                title_col = colidx
                #print(sheet.cell(title_row, title_col).value)
            if cell.value == "Выполнено с начала строительства":
                complete_row = rowidx
                complete_col = colidx
                #print(sheet.cell(complete_row, complete_col).value)
            if cell.value == "Задание на месяц":
                mounth_row = rowidx
                mounth_col = colidx
                #print(sheet.cell(mounth_row, mounth_col).value)
            if cell.value == 'Всего работ':
                stop = rowidx
                stop_rows.append(stop)
            elif cell.value == 'Наименование и марка техники (механизма), оборудования':
                stop = rowidx
                stop_rows.append(stop)
            if cell.value == "Наименование и марка техники (механизма), оборудования":
                row = rowidx
                col = colidx
                equipment.append((row, col))
            if cell.value == "Наименование должностей, профессий":
                row = rowidx
                col = colidx
                human.append((row, col))
            if 'Газокомпрессорная станция' in str(cell.value):
                row_1 = rowidx
                col_1 = colidx
                obj_rows.append(row_1)
                #print(sheet.cell(row_1, col_1).value)
            if 'ВЛ-35 кВ' in str(cell.value):
                row_2 = rowidx
                col_2 = colidx
                #print(sheet.cell(row_2, col_2).value)
                obj_rows.append(row_2)
            if 'Газопровод внешнего транспорта' in str(cell.value):
                row_3 = rowidx
                col_3 = colidx
                obj_rows.append(row_3)
                #print(sheet.cell(row_3, col_3).value)

    dicts = []
    if row_1!=0 and row_2!=0:
        dict_1 = work_dict_gks(sheet, row_1, row_2, id_col)
    elif row_1!=0 and row_3!=0:
        dict_1 = work_dict_gks(sheet, row_1, row_3, id_col)
    else:
        dict_1 = work_dict_gks(sheet, row_1, sheet.nrows, id_col)
    dicts.append(dict_1)

    if row_2!=0 and row_3!=0:
        dict_2 = work_dict_gks(sheet, row_2, row_3, id_col)
    else:
        dict_2 = work_dict_gks(sheet, row_2, sheet.nrows, id_col)
    dicts.append(dict_2)

    if row_3!=0:
        dict_3 = work_dict_gks(sheet, row_3, sheet.nrows, id_col)
        dicts.append(dict_3)

    #print(dicts)

    if stop_rows:
        stop = stop_rows[0]
    else:
        stop = sheet.nrows
    # print(stop)

    params = {'work_row':work_row,
                          'work_name_col':work_name_col,
                          'unit_col':unit_col,
                          'volume_col':volume_col,
                          'start_col':start_col,
                          'start_row':start_row, 
                          'stop_row':stop_row,
                          'stop_col': stop_col,
                          'complete_col':complete_col,
                          'month_col':month_col,
                          'current_col':current_col,
                          'whole_col':whole_col,
                          'days_row':days_row,
                          'days_col':days_col,
                          'months':months,
                          'mounth_col':mounth_col,
                          'title_col':title_col,
                          'comment_col':comment_col,
                          'stop': stop,
                          'id_col': id_col
                        }

    params_res = {     
                        'month_col':month_col,
                        'days_col':days_col,
                        'months':months,
                        'mounth_col':mounth_col,
                        'title_col':title_col,
                        'comment_col':comment_col,
                        }   

    f = {"file_name": filename}
    #print(f)

    #Работы
    work_lst = []
    ROWS = get_rows(sheet, work_name_col, stop)
    for ROW in ROWS: 
#         print(ROW)
        wrk = get_info(sheet, ROW, dicts, dspmk, **params)
        work_lst.append(wrk)
    #print(work_lst)

    #Ресурсы
    resources = equipment+human
    resources_rows = [item[0] for item in sorted(resources)]
    resources_cols = [item[1] for item in sorted(resources)]
    resources_cols
    stops=resources_rows[1:]
    stops.append(sheet.nrows)

    resource_lst=[]
    ROWS = []
    for i, resource_row in enumerate(resources_rows):
        R = range(resource_row, stops[i])
        ROWS.append(R)

    for i, interval in enumerate(ROWS):
        #print(resources_cols[i])
        for ROW in interval:
            r = get_info_res(sheet, ROW, resources_rows[i], resources_cols[i], **params)          
            resource_lst.append(r)
    #print(resource_lst)

    r_lst = [item for item in resource_lst if item['resource_name']!='Наименование и марка техники (механизма), оборудования']
    r_lst = [item for item in r_lst if item['resource_name']!='Наименование должностей, профессий']
    r_lst = [item for item in r_lst if item['resource_name']!='']
    r_lst = [item for item in r_lst if isinstance(item['resource_name'], str)]

    f.update({'work': work_lst})
    f.update({'resource': r_lst})
    print(f)
    
    path_out = 'Z://GPN_KIP//parsed//ТИП//2016//GKS//' + name + '.json'
    with open(path_out, "w", encoding="utf-8") as file:
        json.dump(f, file)


# In[ ]:


# def work_dict_gks(sheet, work_row, id_col, work_name_col):
#     keys = []
#     values = []
#     if 'Наименование этапа строительства' in sheet.col_values(work_name_col):
#         k = 6
#     else:
#         k=4
#     for i in range(work_row+k, sheet.nrows):
#         keys.append(str(sheet.cell_value(i, id_col)))
#         values.append(sheet.cell(i, work_name_col).value)
#     if not '№ п/п' in keys:
#         keys.append('№ п/п')
#     cut = keys.index('№ п/п')
#     keys = [item for item in keys[:cut] if item!='']
#     keys = [item.replace('.0','') for item in keys]
#     values = [item for item in values if item!='' and item!='Наименование этапа строительства']
#     values = [item for item in values[:cut] if item!='']
#     work_dict = dict(zip(keys, values))
#     #print(work_dict)
#     return work_dict

