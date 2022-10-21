#!/usr/bin/env python
# coding: utf-8

# ### Вторая версия парсера, изменения - сохранение объемов работ и добавление обработки xlsx 

# Также высылаю расширенную версию json. Так как выяснили, что физ. объёмы полей, связанных с завершённостью работы по проекту в целом и в месяц, а также остатков по проекту, всё-таки важны и не всегда они пересчитываются из процентов. Теперь такие поля как complite_state, current_remain, whole_remain и mounth_complite имеют две вариации - в процентах и в физ. объёме. Заполняем по факту наличия в данных, если есть оба - заполняем оба варианта, если нет - заполняем что есть.

# In[ ]:


import xlrd
from xlrd import xldate_as_datetime
import re
import pandas as pd
from datetime import datetime
import os
from tqdm import tqdm_notebook
from collections import Counter
import numpy as np
import win32com.client as win32
import pyexcel as p
import json
import traceback
import openpyxl
import msoffcrypto
import xlsxwriter


# In[ ]:


#Не вывожу ошибки
import warnings
warnings.filterwarnings('ignore')


# In[ ]:


#Меняем директорию для удобства обработки
os.chdir(r"C:\Users\Илья\Desktop\Work\ГПН КИП-2\My parsers")


# #### Вспомогательные функции и переменные

# In[ ]:


M = {'Январь':1,
     'Февраль':2,
     'Март':3,
     'Апрель':4,
     'Май':5,
     'Июнь':6,
     'Июль':7,
     'Август':8,
     'Сентябрь':9,
     'Октябрь':10,
     'Ноябрь':11, 
     'Декабрь':12}

M_={numb:name for name,numb in M.items()}

words_type_table={("наименование работ","план"):"work",
                  ("наименование и марка",):("equipment","None"),
                  ("наименование должност",):("human","None"),
                  ("наименование субподрядн","наименование должност"):("human","subcontracting"),
                  ("наименование субподрядн","наименование и марка"):("equipment","subcontracting")
                 }

is_not_None=lambda v:v not in [None,"",str(None)]

    
def val_is_None(row_values,num_val):
    if len(row_values)>num_val:
        is_None=row_values[num_val] in [None,"",str(None)]
    else:
        is_None=True
    return is_None


def check_type_table(cols):
    line_cols=" ".join(cols).lower()
#     print(line_cols)
    
    for words in words_type_table:
        if all([(w in line_cols) for w in list(words)]):
#             print(f"Тип определен как {words_type_table[words]}")
            return words_type_table[words]
#     print(f"Тип определен как None")
    

def create_cols(sh,start_table_rx,last_head_rx,type_file):
    if type_file=="xls":
        names=list(zip(*[[str(cell.value) for cell in sh.row(rx)] for rx in range(start_table_rx,last_head_rx+1)]))
    if type_file=="xlsx":
        names=list(zip(*[[str(cell.value) for cell in sh[rx]] for rx in range(start_table_rx,last_head_rx+1)]))
    
    name_to_line=lambda list_names:" ".join([n for n in list_names if n!=str(None)])
    return list(map(name_to_line,names))

#Органичения - если в индексе буквы индексом считатся не будет
def val_is_index(val):
    return ((re.search(r"^[0-9]+.[0-9]+",str(val).strip(".").strip())!=None)|(str(val).isdigit()))&("resource" not in str(val))

# 'work id': '2001-02-03 00:00:00' заплатка для обработки
def work_id_to_format(work_id):
    if re.search(r"[0-9]{4}-[0-9]{2}-[0-9]{2}",str(work_id))!=None:
        parts_id=re.findall(r"[0-9]{4}-[0-9]{2}-[0-9]{2}",str(work_id))[0].split("-")
        print(parts_id)
        return f"{parts_id[2].lstrip('0')}.{parts_id[1].lstrip('0')}.{parts_id[0][-1]}"
    return work_id

#Запуск заплатки для обработки 
def ind_preprocess(ind):
    ind=work_id_to_format(ind)
    ind=re.sub(r"[^0-9.]","",str(ind).replace(".0","")).strip(".")
    if len(ind)>20:
        ind=str(ind)[:20]
    return ind

#Сокращаем индекс с конца чтобы добавить все высокоуровневые работы, которые указаны, частью которых является эта работа
def get_high_level_works(hierar_i,high_level_works):
    hl_works=[]
    while len(hierar_i)>=1: #Для len(hierar_i)==1 не может быть работ высшего уровня
        hierar_i=".".join(hierar_i.split(".")[:-1])
        if hierar_i in high_level_works:
            hl_works.append(f"{hierar_i} {high_level_works[hierar_i]}")
    return hl_works

def all_vals_int(vals): #,rx_is_first #Модель удалит строку если в ней будут значения 1,2,3,4 например
    if "Примечание" in vals:
        vals.remove("Примечание")
    try:
        if all([str(v).replace(".0","") in list(map(str,list(range(25)))) for v in vals]): #Все значения инт, проверяем
            int_vals=[int(str(v).replace(".0","")) for v in vals]
            if len(int_vals)>=3: #Не понятно как выбирать это число
                if sorted(int_vals)==int_vals: #Проверяем что числа идут по неубыванию:
                    differences=[v2-v1 for v1,v2 in list(zip(int_vals,int_vals[1:]))] #Разница между соседними числами должна быть 1 (0 нельзя,нужные строки попадут)
                    if all(list(map(lambda d:d==1,differences))): 
                        print(f"Нашлась строка числовых столбцов {int_vals}")
                        return True
    except:
        pass

def get_index(last_indexes,row_values):
    val=row_values[0]
    row_values_not_none=[str(v) for v in row_values]
    val_not_none=row_values_not_none[0]
    
    if val_is_index(val):
        if len(row_values_not_none)==2:
            type_index="high_level"
        else:
            type_index="low_level"
        return ind_preprocess(val),type_index
    
    if len(row_values)>1: 
        if val_is_index(row_values[1]): 
            if len(row_values_not_none)==2: #Если индекс во второй строке и всего 2 значения - работа высокоуровневая
                type_index="high_level"
            else:
                type_index="low_level"
            return ind_preprocess(row_values[1]),type_index
    

#     print(f"Значение {val} не распознается как индекc")
    #В случае отсутствия индекса в первых двух значениях считаем что его нет и генерируем из предидущего 
    if len(last_indexes)==0:
        return ("1","high_level")
    else:
        last_index,type_index=last_indexes[-1]
        if type_index=="high_level":
            return (ind_preprocess(last_index)+".1","low_level") #если верхнеуровневый - добавляем .1, 
        
        if type_index=="low_level":
            last_number=str(last_index).split(".")[-1]
            previous_part=str(last_index).split(".")[:-1]
            
            new_index=previous_part+[str(int(last_number)+1)] #Если низкоуровневый - добавляем 1 к последнему числу
            new_index=".".join(new_index)
            return (ind_preprocess(new_index),"low_level")
            
            
#Для файлов СМГ 03 НГСК в высокоуровневых работах в конце строки появляется слово "план" убираем 
def del_words_end_line(value_list):
    row_values_not_none=[str(v) for v in value_list if (v!="")&(v!=None)]
    if len(row_values_not_none)>0:
        if (val_is_index(value_list[0])|val_is_index(value_list[1]))&(len(row_values_not_none)==3): #3 - индекс, значение, план
            if "план" in value_list:
                value_list[value_list.index("план")]=""

    return value_list

#Нужные признаки для json кроме work title и work id
needed_vals_for_json=[
'measurements',
 'amount',
 'start_date_plan',
 'start_date_estimate',
 'start_date_fact',
 'stop_date_plan',
 'stop_date_estimate',
 'stop_date_fact',
 'complite_state_plan',
 'complite_state_fact',
 'whole_remain_value',
 'current_complete_perc',
 'mounth_complite_plan',
 'mounth_complite_fact',
] #'comments'
    
def have_values_for_json(row,new_name_cols):
    have_values=False
    row_values=[cell.value for cell in row]
    #Проверяем заполненно ли хоть одно значение из нужных для json
    if any([(val not in [None,"",str(None)]) for col,val in zip(new_name_cols,row_values) if col in needed_vals_for_json]):
        have_values=True
    return have_values

def row_include_work(row,prev_row,next_row,new_name_cols):
    include_work=False
    row_values=[cell.value for cell in row]
#     print("Проверяем включает ли работы строка",row_values)
    
    #Если в строке есть слово план, следующая строка существует (не None) и есть слово факт
    if next_row!=None:
        next_row_values=[cell.value for cell in next_row]
        if ("план" in row_values)&("факт" in next_row_values):
            include_work=True


    #Если в строке есть слово факт, предидушщая строка существует (не None) и в ней есть слово план
    if prev_row!=None:
        prev_row_values=[cell.value for cell in prev_row]
        if ("факт" in row_values)&("план" in prev_row_values):
            include_work=True
    
    #Функция проверяет строку на наличие работы вне зависимости от наличия слов "план" и "факт"
    #Если у строки есть индекс и второе непустое значение + заполнено хоть 1 значение из нужных для json
    #И следующая строка равна по длинее этой, и не содержит индекса в первых трех значениях
    #Считаем что это пара строк план-факт
    def find_work_row_without_plan_fact(first_row,second_row,new_name_cols):
        include_work=False
        first_row_values=[cell.value for cell in first_row]
        second_row_values=[cell.value for cell in second_row]

        if (len(first_row_values)>=3) & (len(second_row_values)>=3) & (len(new_name_cols)>=3):
            first_row_have_index=val_is_index(first_row_values[0])|val_is_index(first_row_values[1]) #Проверяем наличие индекса в первых двух значениях
            second_row_empty_three_first_values=val_is_None(second_row_values,1)&val_is_None(second_row_values,2)&val_is_None(second_row_values,3)
            
            if first_row_have_index&second_row_empty_three_first_values&have_values_for_json(first_row,new_name_cols):
                include_work=True
        return include_work
            
    if next_row!=None:
        if find_work_row_without_plan_fact(first_row=row,second_row=next_row,new_name_cols=new_name_cols):
            include_work=True
        
    if prev_row!=None:
        if find_work_row_without_plan_fact(first_row=prev_row,second_row=row,new_name_cols=new_name_cols):
            include_work=True
        
    return include_work

#Проверяем стоит ли сохранять информацию из строки
def is_valide(row,prev_row,next_row,type_table,new_name_cols):
    needed_row=False
    row_values=[cell.value for cell in row]
    row_values_not_none=[str(v) for v in row_values if (v!="")&(v!=None)]
    
    first_val=row_values[0]
    if first_val not in ["",None]: #
        needed_row=True
    else: 

        #Если второе значение индекс, строку тоже обрабатываем
        if (len(row_values_not_none)>0)&(len(row_values)>1):
            if val_is_index(str(row_values[1])): #|("gpn" in row_values_not_none[0])
                needed_row=True
        
        #Если строка включает данные о работах - берем ее
        if row_include_work(row,prev_row,next_row,new_name_cols):
            needed_row=True
        
        #Для ресурсов обрабатываем все строки
        if type_table!="work":
            needed_row=True
    
    #Если первые 3 значения цифры 1,2,3 или 2,3,4 - строка числовая, пропускать ее нельзя
    try:
        three_vals=list(map(int,row_values[:3]))
        if three_vals==[1,2,3]:
            needed_row=False
    except:
        pass
    
    try:
        three_vals=list(map(int,row_values[1:4]))
        if three_vals==[2,3,4]:
            needed_row=False
    except:
        pass
    
    return needed_row

#Достаем предидущую и следующую строку
def get_next_prev_rows(rx,end_table_rx,last_head_rx):
    prev_row,next_row=None,None
    if rx!=(last_head_rx+1):
        if type_file=="xls":
            prev_row=sh.row(rx-1)
            
        if type_file=="xlsx":
            prev_row=sh[rx-1]
    
    if rx!=end_table_rx: #Проверяем что значение не последнее
        if type_file=="xls":
            next_row=sh.row(rx+1)
        
        if type_file=="xlsx":
            next_row=sh[rx+1]
    return prev_row,next_row

#Вытаскиваем последний высокоуровневый индекс и привязываем к нему ресурс
def get_last_high_ind(last_indexes):
    last_high_indexes=[i for i in last_indexes if i[1]=="high_level"]
    if len(last_high_indexes)==0:
        return "0"
    else:
        return last_high_indexes[-1][0]


def create_vals(sh,last_head_rx,end_table_rx,type_file,type_table,new_name_cols):
    high_level_works={}
    all_row_vals=[]
    last_indexes=[]
    
    for rx in range(last_head_rx+1,end_table_rx+1):
        
        if type_file=="xls":
            row=sh.row(rx)
            
        if type_file=="xlsx":
            row=sh[rx]
        
        #Достали следующую и предидущую строки
        prev_row,next_row=get_next_prev_rows(rx,end_table_rx,last_head_rx)
        print("row_values_1",[cell.value for cell in row])
        if is_valide(row,prev_row,next_row,type_table,new_name_cols): #Отсеиваем строки с пустым первым значением + без объемов работ
            row_values=[cell.value for cell in row] #Значения не в строчном виде для сохранения и дальнейшего извлечения дат 
            row_values=del_words_end_line(row_values) #Заплатка - удаляем слов план из конца высокоуровневой строки чтобы ловилась 
            row_values_not_none=[str(v) for v in row_values if (v!="")&(v!=None)]
            row_line=" ".join(row_values_not_none)
            
            
            #тестирую убрал (not all_vals_int(row_values_not_none)) вообще
            if (type_table=="work"): #,rx_is_first решил ослабить условие
                
                #Если индекс пустой - считаем что это ресурс и генерируем для него индекс
                if row_values[0] in ["",None]: #!
                    #Вытаскиваем последний высокий индекс
                    last_high_i=get_last_high_ind(last_indexes)
                    row_values[0]=f"{last_high_i}_resource" #Заполняем первое значение, индекс - привязка к работе+resource
                    
                    #Обновляем строки с учетом обновления первого значения
                    row_values_not_none=[str(v) for v in row_values if (v!="")&(v!=None)]
                    row_line=" ".join(row_values_not_none)
                    
#                   if val_is_None(row_values,2)&val_is_None(row_values,3): #Если 3 или 4 значение не пустые - не ресурс,заплатка для урм-шинг  
                    #ЗАКОМЕНЧЕНО - считаем что все строки с пустыми индексами это ресурсы
                    #Вытаскиваем индекс строки (если его нет - генерируем) и его тип (низкоуровневый или высокоуровневый)
    #               new_index,type_index=get_index(last_indexes,row_values)

                
                
                
#                 print("row_values_2",row_values,len(row_values))
                
                #! убрано &(len(row_values_not_none)==2) для теста
#                 print("rule 1",val_is_index(row_values[0])&((not val_is_None(row_values,1))|(not val_is_None(row_values,2)))&(not row_include_work(row,prev_row,next_row,new_name_cols))) # &((not val_is_None(row_values,1))|(not val_is_None(row_values,2)))
                if val_is_index(row_values[0])&((not val_is_None(row_values,1))|(not val_is_None(row_values,2)))&(not row_include_work(row,prev_row,next_row,new_name_cols)): #2 значения + второе не пустое - индекс считаем высокоуровневым
#                     print(1)
                    high_level_works[ind_preprocess(row_values_not_none[0])]=row_values_not_none[1]
                    last_indexes.append((ind_preprocess(row_values_not_none[0]),"high_level"))
                    
                    
                #Если индекс на втором месте, третья не пустая, и длинна 2 - считаем высокоуровневой работой
                #! убрано &(len(row_values_not_none)==2) для теста
                elif (val_is_index(row_values[1]))&((not val_is_None(row_values,2))|(not val_is_None(row_values,3)))&(not row_include_work(row,prev_row,next_row,new_name_cols)): #!  
                    high_level_works[ind_preprocess(row_values[1])]=row_values[2]
                    last_indexes.append((ind_preprocess(row_values[1]),"high_level"))
                
                #!Если индекс определен и есть работы в строке - прибавляем высокоуровневые работы
                elif (val_is_index(row_values[0]))&(row_include_work(row,prev_row,next_row,new_name_cols)):  #Только если индекс на первом месте добавляем работы высокого уровня
                    
                    #Вытаскиваем все работы высшего уровня к которым относится работа в строке
                    hl_works=get_high_level_works(ind_preprocess(row_values[0]),high_level_works)
                    #Добавляем к строке все работы высшего уровня
                    row_values+=hl_works
                    all_row_vals.append(row_values)
                    last_indexes.append((ind_preprocess(row_values[0]),"low_level"))
                
                
# Закоменчено, так как следующие условие тоже добавляет строки ресурсов                
#                 elif row_values[0]=="resource":
#                     all_row_vals.append(row_values)
                    
                elif row_include_work(row,prev_row,next_row,new_name_cols): #Если есть работы в строке, но первое значение не индекс (для факта)
                    all_row_vals.append(row_values)
                else:
#                     print(f"Строка не сохранена {row_values}")
                    pass
                
            else: #Добавляем значения для ресурсов
                if not all_vals_int(row_values_not_none):
                    all_row_vals.append(row_values)
#     print("all_row_vals",all_row_vals)            
#     print("high_level_works",high_level_works)
    return all_row_vals #!Неравная длинна, возможно не равна длинне колонок


for_rename_cols={
   ( "наименование работ",):"work title",
    ("п/п",):'work id',
    ("ед.","изм"):"measurements",
    ("всего",):"amount", #"кол-во",
    ("начал",'работ',"план"):"start_date_plan",
    ("начал",'работ',"ожид"):"start_date_estimate",
    ("начал",'работ',"факт"):"start_date_fact",
    ("оконч","план"):"stop_date_plan",
    ("оконч","ожид"):"stop_date_estimate",
    ("оконч","факт"):"stop_date_fact",
    ("выполн","нач","план"):"complite_state_plan",
    ("выполн","нач","факт"):"complite_state_fact",
    ("остат",):"whole_remain_value",
#     ("общий","%"):,
    ("%","выполн","меся"):"current_complete_perc", #,
    ("месяц","план"):"mounth_complite_plan",
    ("задан","факт"):"mounth_complite_fact",
#     ("%","выполн","план"):"mounth_complite_fact",
#     ("задан","план"):'mounth_complite_plan', #+,"месяц"
#     ("задан","факт"):'mounth_complite_fact', #+,"месяц"
    tuple([m.lower() for m in list(M)]):1,
    ("примечан",):"comments"
}


#Функция для приведения названия колонок к красивому виду, добавляет фразу которая должна быть в колонке
def get_needed_phrase(col,markers_cols):
    markers_in_col=[w for w in markers_cols if w in str(col).lower()]
    if (len(markers_in_col)>0)&(all([
        (col not in markers_cols) for col in [col,col.strip().strip("\n"),col.strip().strip("\n").lower()]])):
        marker=markers_in_col[0]
        return str(col).lower().replace(marker,"").strip().replace("\n"," ")

#1. После слов план и файт не должно быть ничего, чистим все что после
def clear_plan_fact(line):
    if re.search(r".*факт|.*план",str(line).lower())!=None:
#         line=re.findall(r".*факт|.*план",str(line).lower())[0]
        line=str(line).lower().replace("на дату","").strip() #мешает
        line=re.sub(r" +",' ',line)
        line=line.replace("ожид/","").replace("ожид/","").strip() #удаляем /факт для столбца ожид/факт 
    return line

def rename_cols(table,type_table,for_rename_cols=for_rename_cols):
    
    
#     print(f"Старые названия колонок {table.columns.to_list()}")
    if type_table=="work":
        #Чищу колонки план и факт от окончаний
        table.columns=[clear_plan_fact(c) for c in table.columns]
        print(table.columns)
        
        #Переименовываю колонки c план, факт и ожид к нужному виду
        markers_cols=["план","ожид","факт"]
        is_marker_in_col=lambda col:any([w in str(col).lower() for w in markers_cols])
        get_markers_in_col=lambda col:[w for w in markers_cols if w in str(col).lower()]

        phrase_for_append=None
        new_columns=[]
        for col in table.columns:
            if is_marker_in_col(col):
                if (get_needed_phrase(col,markers_cols) is not None):
                    phrase_for_append=get_needed_phrase(col,markers_cols)
                #Как для строки с фразой для добавления так и для обычной строки добавляем имя единнообразно
                new_columns.append(phrase_for_append+" "+str(get_markers_in_col(col)[0]))
            else:
                new_columns.append(col.replace("\n"," "))
        table.columns=new_columns
        
        
    if type(type_table)==tuple:
        first_part_type=type_table[0]
        if first_part_type=="equipment":
            for_rename_cols={("наименован","марк"):"resource_name",
                       ("п/п",):"resource_id",
                       ("примеч",):'comments'
                      }
        elif first_part_type=="human":
            for_rename_cols={("должност",):"resource_name",
                       ("п/п",):"resource_id",
                       ("примеч",):'comments'
                      }
        else:
            pass
#             print(f"Тип таблицы не определен для {table}")
    
    if (type_table=="work")|(type(type_table)==tuple):
        #Переименовываю все колонки к виду нужному для сохранения работ
        for col in table.columns:
            #Ищу и удаляю месяц
            col_only_text=re.sub("[^А-я]+","",str(col))
            if (col_only_text in M)|(col_only_text.capitalize() in M):
                table.rename({col:re.sub("[^0-9.]+","",str(col))},axis=1,inplace=True)
            
            for words in for_rename_cols:
                if all([w in str(col).lower() for w in words]):
                    table.rename({col:for_rename_cols[words]},axis=1,inplace=True)
            
#     print(f"Новые названия колонок {table.columns.to_list()}")
    
def val_to_float(val,is_error_to_None=False):
    new_val=re.sub(r"[^0-9.]+","",str(val).replace(",",".").strip("."))
    try:
        return round(float(new_val),3)
    except:
        if is_error_to_None:
            return None
        else:
            return val

def to_new_format(date):
    data_pattern=r"[0-9]{4}-[0-9]{2}-[0-9]{2}"
    if re.search(data_pattern,str(date))!=None:
        data=re.findall(data_pattern,str(date))[0]
        return f"{data.split('-')[2]}.{data.split('-')[1]}.{data.split('-')[0]}"
    
    if type(date)==type(datetime(2000, 1, 1)):
        append_zero:lambda val: "0"+str(val) if len(str(val))==1 else str(val)
        return f"{append_zero(date.day)}.{append_zero(date.month)}.{date.year}"            
    return date

def date_preproc(val_date,book):
    try:
        return to_new_format(xldate_as_datetime(val_date,book.datemode))
    except:
        first_var_date=val_date
        
        val_date=re.sub(r"[^0-9.]+","",str(val_date)).strip(".")
        try:
            return to_new_format(xldate_as_datetime(val_date,book.datemode))
        except:
            
            #Работает для xlsx файлов (заплатка, может упасть)
            if val_date[:8].isdigit():
                try:
                    return to_new_format(str(pd.to_datetime(val_date)))
                except:
                    pass 
            
            
            parts_date=val_date.split(".")
            if len(parts_date)==3:
                if len(parts_date[0])==4: #Предполагаю что год первый в дате
                    try:
                        return to_new_format(datetime(int(parts_date[2]), int(date_parts[1]), int(date_parts[0])))
                    except:
                        pass
                elif len(parts_date[0])==2: #Предполагаю что день первый в дате
                    try:
                        return to_new_format(datetime(int(parts_date[0]), int(date_parts[1]), int(date_parts[2])))
                    except:
                        pass
            
            return to_new_format(first_var_date)

# #Проверяем значение на наличие даты, если 14 цифр - считаем что это дата (8 дата + 6 время)
#     '20170101000000', '2017-01-02 00:00:00'
def is_date(val):
    return len(re.sub('[^0-9]','',val))==14

def date_prepros(date):
    if len(date)==len('2017-01-01 00:00:00'):
        return date
    else:
        numerical_date=re.sub('[^0-9]','',str(date))
        return f"{numerical_date[:4]}-{numerical_date[4:6]}-{numerical_date[6:8]} 00:00:00" #! Только для ггггммдд

#Старый вариант
def day_prepros(day):
    day=re.sub(r"[^\w.]","",str(day).replace(".0",""))
#     print("day first",day)
    #Убираем месяц
    day=re.split(r'[ .,/\-_]+', day)[0]
    
#     try:
#         day=str(pd.to_datetime(str(day)).day)
#     except:
#         pass
    
    
    return day


    
def get_first_not_None(series):
    vals=list(series)
    not_None=lambda val:val not in [None,""]
    if any(map(not_None,vals)):
        return [v for v in vals if not_None(v)][0]
    else:
        return ""
#Органичения - если в индексе буквы индексом считатся не будет
def is_numerical(val):
    return (re.search(r"^[0-9]+.[0-9]+",str(val).strip("."))!=None)|(str(val).isdigit())

def calculating_mounth_plan_fact(progress): 
    calculated={"plan":[],"fact":[]}
    for day_values in progress:
        for day,values in day_values.items():
            for plan_or_fact,val in values.items():
                if is_numerical(val):
                    calculated[plan_or_fact]=calculated[plan_or_fact]+[val_to_float(val)]
    return sum(calculated['plan']),sum(calculated['fact'])

#Достаем первое значение если переменная список, если нет - возвращаем переменную без изменений
def get_first_val(val):
    if type(val)==list:
        if len(val)>0:
            return val[0]
    else:
        return val
    
#Функция вытаскивает первое значение, если данные серия или список
def preproc_if_series(val,type_val=None):
    if (type(val)!=type(pd.Series()))&(type(val)!=list):
        return val
    
    else:
        val_list=list(val)

        if (len(val_list)==1)|(type_val!="progress"): #Только для progress допускается несколько значений 
            return str(val_list[0])
        #Обрабатываем если скрипт пытается передать несколько значений в прогресс
        elif (len(val_list)>1)&(type_val=="progress"):
            not_None_vals=[v for v in val_list if (v!=None)&(v!=str(None))&(str(v)!="nan")]
            
            #Выкидываем первое значение если список не пустой, если пустой - None
            if len(not_None_vals)>1: 
                return str(not_None_vals[0])
            else:
                return None
        else:
            return val_list

#Тут построено на row[ind][0], возможно для xlsx это не работает
def append_dict_from_rows_work(row,file_sheet_json):
    work={
        'work title': None, #+
        'work id': None, #+
        'upper works': [],#+
        'measurements': None, #+
        'amount': None, #+
        'work_data': {
        'start_date': {'plan': None, 
                       'estimate': None,
                       'fact': None},#+
            
        'stop_date': {'plan': None,
        'estimate': None,
        'fact': None}, #'date in format dd.mm.yy' #+
        'complite_state_perc': {'plan': None, 'fact': None}, #'in %' #+
        'complite_state_value': {'plan': None, 'fact': None}, #+
        'current_remain_perc': None, #'in %' #+
        'current_remain_value': None, #'in %' #-
        'whole_remain_perc': None,#'in %' #- общий % выполнения -1 +
        'whole_remain_value': None,#+
        'mounth_complite_value': {'plan': None, 'fact': None}, #+
        'mounth_complite_perc': {'plan': None, 'fact': None}, #+'in %'
        'progress': [], #array {'day_id': {'plan': 'value', 'fact': 'value'}
        'comments': 'comment text' #+
        }
    }
          
    words=["plan","estimate","fact"]
    
#     print("Добавляется в work строка",row)
    
    #Тут упадет, если колонки не все строки
    for ind in row.index:
        if ind in list(work['work_data']):
            if ind =='whole_remain_value': 
                row[ind]=val_to_float(preproc_if_series(get_first_val(row[ind][0])))
                print(row[ind])
                
            elif ind=='comments':
#                 print("comments_value",row[ind])
                row[ind]=str(get_first_not_None(row[ind])) #Предидущее row[0]
            else:
                pass
            work["work_data"][ind]=preproc_if_series(row[ind])
        elif ind in list(work):
            if ind=='work id':
                row[ind][0]=ind_preprocess(row[ind][0])
            work[ind]=str(row[ind][0]).strip(" ")

                    
        elif any([w in ind for w in words]):
            first_part,second_part="_".join(ind.split("_")[:-1]),ind.split("_")[-1]
            if first_part in ["complite_state","mounth_complite"]:
                work['work_data'][first_part+"_value"][second_part]=val_to_float(get_first_val(row[ind][0]))
            elif any([w in ind for w in ["start_date","stop_date"]]):
                work['work_data'][first_part][second_part]=str(date_preproc_1(row[ind][0]))
            else:
                pass
                
        
            
            
        elif is_date(ind):
            work['work_data']['progress']=work['work_data']['progress']+[{date_preproc_1(ind): {'plan': val_to_float(row[ind][0],is_error_to_None=True),
                                                                                                'fact': val_to_float(row[ind][1],is_error_to_None=True)
                                                                                               }
                                                                         }]
        else:
            pass
    
    
    
    #Расчитываем current remain percent and value
    mounth_values=work['work_data']["mounth_complite_value"]
    if is_not_None(mounth_values["fact"])&is_not_None(mounth_values["plan"]):
        try:
            mounth_fact,mounth_plan=mounth_values["fact"],mounth_values["plan"]
            if (mounth_fact<=mounth_plan)&(mounth_plan!=0):
                work["work_data"]["current_remain_perc"]=round((100-mounth_fact*100/mounth_plan),2)
                work["work_data"]['current_remain_value']=round(mounth_plan-mounth_fact,2)
            else:
                work["work_data"]["current_remain_perc"],work["work_data"]['current_remain_value']=0,0
        except:
            pass
                

#Убираем так как возможно правильно что план за месяц не совпадает с суммарным, может быть в этом есть идея  
#     #calculating_mounth_plan_fact
#     progress=work['work_data']['progress']
#     print(progress,work['work_data']['progress'])
#     work['work_data']["mounth_complite_value"]['plan']=calculating_mounth_plan_fact(progress)[0]
#     work['work_data']["mounth_complite_value"]['fact']=calculating_mounth_plan_fact(progress)[1]
            
    if (work["amount"] not in [None,0]):
        #whole_remain to percent
        if work["work_data"]["whole_remain_value"]!=None:
            try:
                whole_remain_=work["work_data"]["whole_remain_value"]
                amount_=work["amount"]
                work["work_data"]["whole_remain_perc"]=round((whole_remain_*100)/val_to_float(amount_),2)
            except:
                pass

    
        #mounth_complite to percent
        for first_part in ['complite_state',"mounth_complite"]:
            for second_part in ["plan","fact"]:
                if work["work_data"][first_part+"_value"][second_part]!=None:
                    try:
                        val1=work["work_data"][first_part+"_value"][second_part]
                        val2=work["amount"]

                        work["work_data"][first_part+"_perc"][second_part]=round((val1*100)/val_to_float(val2),2)
                    except:
                        pass
    
    #Заплатка - Если mounth_complite_value plan==0 current_remain_perc должно быть равно None
    if work["work_data"]['mounth_complite_value']['plan']==0:
        work["work_data"]['current_remain_perc']=None
    
    #Название работ цифровые скрипт не сохраняет
    if (val_is_index(work['work id'])|(work['work title']!=None))&(not str(work['work title']).isdigit()):
        file_sheet_json['work']+=[work]
        #: #&(not (work["amount"] in [None,""])

#         print("Успешно добавлена",work)




def append_resource_from_work_table(row,file_sheet_json):   
    resource={
        'resource_id': None, #'id' #
        'resource_name': None,#'name' #
        'type': None, #'human/equipment/nan' #
        'progress': [], #            {'day_id': [{'plan': 'amount','fact': 'amount'}
        'comments': None} #'comment text' #

    progress_values=[]
    for ind in row.index:
        if ind=='comments':
            row[ind]=str(get_first_not_None(row[ind])) 
        elif ind=='work title':
            resource['resource_name']=row[ind][0] #Так как таблица - работа, resource_name=work title
        elif ind=='work id':
            resource['resource_id']=row[ind][0] #Так как таблица - работа, work id=resource_id
        elif is_date(ind):
            resource['progress']=resource['progress']+[{date_preproc_1(ind): {'plan': val_to_float(row[ind][0],is_error_to_None=True),
                                                                              'fact': val_to_float(row[ind][1],is_error_to_None=True)
                                                                             }}]
            progress_values.append(str(preproc_if_series(row[ind],type_val="progress")))
        else:
            pass
     
    #Сохраняем если заполнено resource_name или любое значение прогресса непустое и есть индекс
    if str(resource["resource_name"]).replace(".","",1).isdigit()==False: #Имя ресурса не может быть числовым
        if ((str(resource["resource_name"]).strip()!='')&(resource["resource_name"]!=None))|(
            any(list(map(lambda v:v not in [None,"",str(None),"nan",["",""],[None,None]],progress_values)))&(val_is_index(resource["resource_id"]))):
            file_sheet_json['resource']+=[resource]      
    


def append_dict_from_rows_recourses(row,file_sheet_json,table_have_plan_fact,type_resource):
    resource={
        'resource_id': None, #'id' #
        'resource_name': None,#'name' #
        'type': type_resource, #'human/equipment/nan' #
        'progress': [], #            {'day_id': [{'plan': 'amount','fact': 'amount'}
        'comments': None} #'comment text' #
    

    
    progress_values=[]#Найденные значения погресса
    for ind in row.index:
        if ind in list(resource):
            resource[ind]=get_first_val(preproc_if_series(row[ind]))
        
        elif is_date(ind):
            if not table_have_plan_fact:
                resource['progress']=resource['progress']+[{date_preproc_1(ind):preproc_if_series(row[ind],type_val="progress")}]
                progress_values.append(str(preproc_if_series(ind,type_val="progress")))
            else:
                resource['progress']=resource['progress']+[{date_preproc_1(ind): {'plan': val_to_float(row[ind][0],is_error_to_None=True),
                                                                                  'fact': val_to_float(row[ind][1],is_error_to_None=True)}}]

        elif (day_prepros(ind) in list(map(str,list(range(32))))):
            if not table_have_plan_fact:
                resource['progress']=resource['progress']+[{day_prepros(ind):preproc_if_series(row[ind],type_val="progress")}]
                progress_values.append(str(preproc_if_series(row[ind],type_val="progress")))
            else:
                resource['progress']=resource['progress']+[{ind: {'plan': val_to_float(row[ind][0],is_error_to_None=True),
                                                                  'fact': val_to_float(row[ind][1],is_error_to_None=True)
                                                                 }
                                                           }]
                
        else:
#             print(f"Не удалось добавить в json информацию из {str(ind)}")
            pass
    
 
    #is c
    #Сохраняем если заполнено resource_name или любое значение прогресса непустое и есть индекс
    if str(resource["resource_name"]).replace(".","",1).isdigit()==False: #Имя ресурса не может быть числовым
        if ((str(resource["resource_name"]).strip()!='')&(resource["resource_name"]!=None)):
#         |(any(list(map(lambda v:v not in [None,"",str(None),"nan",["",""],[None,None]],progress_values)))&(val_is_index(resource["resource_id"]))):
            file_sheet_json['resource']+=[resource]              

            
        
#Удаляю повторение названий колонок (одно и то же написано дважды)       
def del_double_name_cols(col):
    parts=col.strip().split(" ")
    print(parts)
    if len(parts)==4:
        if parts[:2]==parts[2:4]:
            return " ".join(parts[:2])
    return col
    
def table_preprocessing(table,file_sheet_json,type_table):
    print(f"Обрабатывается таблица типа {type_table}")
    #переименовываем колонки, в зависимости от типа таблицы
    print("table",table)
    rename_cols(table,type_table=type_table)
    
    if type_table=="work":
        #cобираю высокоуровневые работы в одну колонку
        table['upper works']=pd.Series(
            zip(*[table[col] for col in table.columns if "high work" in col]),table.index).apply(lambda L:[
            v for v in L if (v!=None)&(v!="")])
        
        #схлопываем индексы (тут могут возникать проблемы)
        table['for_groupby']=table.index.map(lambda ind:ind if ind%2==0 else ind-1)
        #Если количество нечетное - что то пошло не так, обрезаем последнюю строкуц
        if table.shape[0]%2!=0:
            table=table[:-1]
        
        
#         if type_table=="work":
#             try:
#                 column_days=[c for c in table.columns if "дни" in str(c).lower()][0]
#                 #План должен стоять на четном месте, факт на нечетном. Удаляем по одной несоответстсвующие строки
                
#                 table=table[((table[column_days]=="план")&(table.index%2==0))|((table[column_days]=="факт")&(table.index%2==1))]
#             except:
#                 pass
        
        table=table.groupby("for_groupby").agg(list)
#         print('new_table',table)
#         print('new_table_columns',table.columns.tolist())
        
        #Удаляю все дубликаты колонок, оставляю последние (актуально для файлов ДСМПК с повторяющимися колонками дней)
        table.columns=[str(c).strip(" ").replace(".0","") for c in table.columns]
        
        #Удаляю дубликаты  
        table.columns=[del_double_name_cols(col) for col in table.columns]
        
        print("Колонки финальные названия",list(table.columns))
#         table = table.loc[:,~table.columns.duplicated(keep="last")] #Считаем что последние значения наиболее актуальные
        
#         # Заплатка - переименовываю 30 день в 31 если 30 дней 2
#         cols=table.columns.to_list()
#         if cols.count("30")==2:
#             try:
#                 last_ind=cols.index("30")+1
#                 while cols[last_ind]!="30":
#                     last_ind+=1
#                 cols[last_ind]="31"  
#                 table.columns=cols
                
#             except:
#                 pass
        
        #добавляю все работы из table к file_sheet_json, и ресурсы по отдельному алгоритму
        for i,row in table.iterrows():
#             print(row.values)
            if "resource" in str(get_first_val(preproc_if_series(row['work id']))):
                append_resource_from_work_table(row,file_sheet_json) #если resource в индексе - добавляем к ресурсам
            else:
                append_dict_from_rows_work(row,file_sheet_json) #если нет - к работам
            


        
    
    elif type(type_table)==tuple: #Таблица ресурс
        table_have_plan_fact=False
        
        cols_line=" ".join(table.columns.tolist())
        print("cols_line_1",cols_line)
        for c in table.columns:
            print(c,set(table[c]))
        #Проверяем есть ли столбец с план-факт в значениях
        pal_fact_in_col=any([all([w in set(table[c]) for w in ["факт","план"]]) for c in table.columns])
        print("pal_fact_in_col",pal_fact_in_col)
        
        #Группирую строки так как там есть план факт
        if (re.search(r"дни",cols_line.lower())!=None)|(pal_fact_in_col): #Если в колонке есть столбик со словом дни - план факт там должен быть
            table_have_plan_fact=True
            
            #схлопываем индексы (тут могут возникать проблемы)
            table['for_groupby']=table.index.map(lambda ind:ind if ind%2==0 else ind-1)
            
            #Если количество нечетное - что то пошло не так, обрезаем последнюю строкуц
            if table.shape[0]%2!=0:
                table=table[:-1]
            
            table=table.groupby("for_groupby").agg(list)
            print("table_1",table)
            
        
        first_part_type=type_table[0]
        
        #Удаляю все дубликаты колонок, оставляю последние (актуально для файлов ДСМПК с повторяющимися колонками дней)
        table.columns=[str(c).strip(" ").replace(".0","") for c in table.columns]
#         table = table.loc[:,~table.columns.duplicated(keep="last")] #Считаем что последние значения наиболее актуальные
#         print('table_resources',table)
#         print('table_resources_columns',table.columns.tolist())
        
        #Удаляю дубликаты  
        table.columns=[del_double_name_cols(col) for col in table.columns]
    
#         Заплатка - переименовываю 30 день в 31 если 30 дней 2
#         cols=table.columns.to_list()
#         if cols.count("30")==2:
#             try:
#                 last_ind=cols.index("30")+1
#                 while cols[last_ind]!="30":
#                     last_ind+=1
#                 cols[last_ind]="31"  
#                 table.columns=cols
                
#             except:
#                 pass
        
        [append_dict_from_rows_recourses(row,
                                         file_sheet_json,
                                         table_have_plan_fact,
                                         type_resource=first_part_type) for i,row in table.iterrows()]
        
#Переименовываю колонку с план/факт
def rename_plan_fact_col(table,cols):
    for i in range(table.shape[1]):
        col_i=table.columns.tolist()[i]
        strip_all_val=lambda List:list(map(lambda v:str(v).strip(),List))
        if all([w in strip_all_val(list(table[col_i])) for w in ['план','факт']]):
            all_cols_line="".join(cols).lower()
            if (len(cols)-1>=i)&(re.search("дни мес",all_cols_line)==None):
                cols[i]="Дни мес."
            
def table_to_json(file_sheet_json,
                  sh,
                  start_table_rx,
                  end_table_rx,
                  last_head_rx,
                  type_file):
    
    #Creating df
    cols=create_cols(sh,start_table_rx,last_head_rx,type_file)
#     print(f"Извлеченные названия колонок, {cols}")
    
    #Type table
    type_table=check_type_table(cols)
    
    #Сreate new names cols for select rows with work
    empy_table=pd.DataFrame(columns=cols)
    rename_cols(empy_table,type_table=type_table)
    new_name_cols=empy_table.columns.tolist()
#     print("new_name_cols",new_name_cols)
    
    vals=create_vals(sh,last_head_rx,end_table_rx,type_file,type_table,new_name_cols)
    table=pd.DataFrame(data=vals) #!vals и cols не равны
    
    #Нахожу и переименовываю колонку с план/факт в cols
    rename_plan_fact_col(table,cols)
#     print("cols_1",cols)
        
            
            
#     print(table)
    
    if table.shape[0]==0:
        print(f"Получена пустая таблица {start_table_rx,end_table_rx,last_head_rx}")
    else:
        table.columns=cols+[f"high work level {i}" for i in range(table.shape[1]-len(cols))]
        new_cols=table.columns.to_list()
#         print("Смотрю значения 1")
        for c in table.columns:
            print(c,set(table[c]))
        table_preprocessing(table,file_sheet_json,type_table=check_type_table(new_cols))


# In[ ]:


#Проверяем файл уже обрабатывался или нет
def is_file_already_preprocessed(file,sh_name):
    files=pd.read_csv("Successfully preprocessed files.xlsx",index_col="index")
    return files[(files["files"]==file)&(files["sh_name"]==sh_name)].shape[0]!=0


#Добавляем уже обработанный файл в табличку
def append_preprocessed_file(file,sh_name):
    files=pd.read_csv("Successfully preprocessed files.xlsx",index_col="index")
    if files.shape[0]==0:
        files.loc[0]=file
    else:
        last_index=files.index.to_list()[-1]
        files.loc[int(last_index)+1]=[file,sh_name]
    files.to_csv("Successfully preprocessed files.xlsx",index_label="index")


# In[ ]:


#Снятие защиты с защищенного файла xsl
def decrypt_xlsx_files(file,full_path_parsing):
    input_path=full_path_parsing
    output_path=os.path.join("Data for parsing","decrypted files")
    
    #Создаем директорию для сохранения расшифрованного файла если ее еще нет
    isExist = os.path.exists(output_path)
    if not isExist:
        os.makedirs(output_path)

    
    data = msoffcrypto.OfficeFile(open(os.path.join(input_path, file), 'rb'))
    data.load_key(password='VelvetSweatshop')  # 默认密码为'VelvetSweatshop'
    data.decrypt(open(os.path.join(output_path, file), 'wb'))  # 输出无密码保护
    print('finished' + file)

    #Открываем книгу
    _book = xlrd.open_workbook(os.path.join(output_path, file))
    return _book


# In[ ]:


#Ищем месяц который будем парсить
def get_month(file_name):
    #Ищем текстовый месяц в названии файла
    for m in M:
        if (len(m)>4)&(m.lower()[:-1] in file_name.lower()):
            return m.lower()
        if m.lower() in file_name.lower():
            return m.lower()
    
    numbs=re.sub("[^0-9 ,-_]","",file_name)
    numbs=re.sub("[ ,-/_]"," ",numbs)
    
    numbs=re.sub(" +"," ",numbs)
    
    parts=numbs.strip().split(" ")
    print(parts)
    #месяц всегда второй
    try:
        #Ищем год
        if (len(parts[0])==4)|(len(parts[-1])==4): #Предполагаем что минимум две части есть, и год сбоку
            if "20" in str(parts[0]):
                year=0
                month=1
            elif "20" in str(parts[-1]):
                year=-1
                month=-2
            else:
                return str(M_[int(parts[1])]).lower()
            
            return str(M_[int(parts[month])]).lower()
        
        elif (len(parts[0])==2)|(len(parts[-1])==2):
            num_0=int(parts[0])
            num_minus_1=int(parts[-1])
            if num_minus_1 in range(15,18): #Обычно год вконце
                year=-1
                month=-2
                
            elif num_0 in range(15,18):
                year=0
                month=1
            else:
                return str(M_[int(parts[1])]).lower()
            
            return str(M_[int(parts[month])]).lower()
        else:
            return ""
    except:
        try:
            return str(M_[int(parts[1])]).lower()
        except:
            print(f"Не получилось найти месяц в названии {file_name}")
            return "" #При таком выводе обрабатываются и сохраняются все листы
        



# In[ ]:


#1_Мессояха_MСГ_30.01.2015.json
#Создаем имя файла
def get_index_for_filename(file_name,full_path_save):
    indexes_in_folder=[int(file_name.split("_")[0]) for file_name in os.listdir(full_path_save)]
    if len(indexes_in_folder)==0:
        return 0
    else:
        return max(indexes_in_folder)+1

def create_file_name(file_name,path_parsing,full_path_save):
    year=path_parsing.split(r"/")[0] #Определяем год по имени папки с файлом
    day,month="xx","xx"
    second_part_year=year[2:] #Год всегда 4 цифры
    
    numbs=re.sub("[^0-9 .,/\-_]","",file_name)
    numbs=re.sub("[ .,/\-_]"," ",numbs)
    numbs=re.sub(" +"," ",numbs).strip()
    
    parts=numbs.strip().split(" ")
    if len(parts)>3: #2 раза указан год и возможно есть какой то мусор, находим эти 2 раза и обрезаем лишнюю часть
        new_parts=[]
        for p in parts:
            if second_part_year in p:
                if parts.index(p) in [0,2]: #Первые 3 числа - то что нужно
                    parts=parts[:3]
                    break
                if parts.index(p)==1: #считаем что первое число - год
                    parts=parts[1:4]
                    break
                if parts.index(p)>2:
                    parts=parts[parts.index(p)-2:parts.index(p)+1]
                    break
        
    try:
        #В какой то части должен быть год, либо в первой либо в последней
        if (second_part_year in parts[-1])&(second_part_year not in parts[0]):
            #Если в последней значит день-месяц в первых двух
            if len(parts)==3:
                month=parts[1]
                day=parts[0]
            elif len(parts)==2:
                #предполагаем что записан только месяц
                month=parts[0]
                if len(month)>2: #cчитаем что день и месяц слиплись
                    day=month[:-2] #День первый если год последний
                    month=month[-2:]
                    
            else:
                print("Месяц не определяется")
                
        elif (second_part_year in parts[0])&(second_part_year not in parts[-1]): #год в первой части, нет в последней
            #Если в первой значит месяц-день две последние
            if len(parts)==3:
                month=parts[1]
                day=parts[-1]
                
            elif len(parts)==2:
                #предполагаем что записан только месяц
                month=parts[1]
                if len(month)>2: #cчитаем что день и месяц слиплись
                    day=month[2:]
                    month=month[:2]
                    
            else:
                pass
#                 print(f"Месяц не определяется для {file_name}")
        elif (second_part_year in parts[0])&(second_part_year in parts[-1]): #День и год совпадают (год и месяц не могут) для годов 15-17
            day=second_part_year
            if len(parts)==3:
                month=parts[1]
    except:
        pass
#         print(f"Дата не определяется для {file_name}")
    
    #Меняю местами месяц и день если месяц больше 12 а день меньше
    try:
        if (int(month)>12)&(int(day)<=12):
            past_day_month=(day,month)
            day,month=past_day_month[1],past_day_month[0]
    except:
        pass
        
            
    #Создаю имя файла
    day_month_preproc=lambda v:"0"+str(v) if len(str(v))==1 else str(v)
    
    date=f"{day_month_preproc(day)}.{day_month_preproc(month)}.{year}"
    index=get_index_for_filename(file_name,full_path_save)
         
    new_file_name=f"{index}_ТИП_{date}"
    for ob in ["ВЛ","ГТЭС"]:
        if ob in file_name:
            new_file_name+=f"_{ob}"
    return new_file_name


# In[ ]:


#Функция поиска имени объекта в файле
def get_object(book,sheets,month_for_parsing,type_file):
    #Ставим нужный sh_name на первое место
    for sh_name in sheets:
        if (re.search(month_for_parsing,sh_name.lower())!=None)&(re.search("правила",str(sh_name).lower())==None)&(re.search("обр",str(sh_name).lower())==None): 
            sheets=[sh_name]+sheets
            break
    
    #Ищем object_,object_2 - первый объектв в шапке, второй в таблице работ
    object_=None
    object_2=None
            
    #Ищем имя объекта последовательно во всех листах, начиная с нужного 
    for sh_name in sheets:
        if re.search("правила",str(sh_name).lower())==None: #Не интересуют листы правила заполнения/составления МСГ
            
            #Файл для сохранения ресурсов
            if type_file=="xls":
                sh = book.sheet_by_name(sh_name)
            if type_file=="xlsx":
                sh = book[sh_name]

            if type_file=="xls":
                num_rows = sh.nrows
            if type_file=="xlsx":
                num_rows = sh.max_row
            
            
                
            for rx in range(num_rows):
                
                if type_file=="xls":
                    row=sh.row(rx)

                if type_file=="xlsx":
                    rx+=1 #Начинается с единицы
                    row=sh[rx]
                row_value=[str(cell.value) for cell in row if str(cell.value) not in [None,"","nan",str(None)]]
                row_line=" ".join(row_value)
                row_value_not_None=[v for v in row_value if is_not_None(v)]
                
                
                if ("по объекту" in row_line):
                    object_=row_line.split("по объекту")[1].strip(":").strip(" ")
                
                try:
                    if list(map(lambda val:str(int(val)),row_value_not_None[:4]))==['1','2','3','4']:
                        new_row=sh[rx+1]
                        new_row_value=[str(cell.value) for cell in new_row if str(cell.value) not in [None,"","nan",str(None)]]
                        new_row_value_not_None=[v for v in new_row_value if is_not_None(v)]
                        if (len(new_row_value_not_None)==1):
                            object_2=new_row_value_not_None[0]
                    
                            if is_not_None(object_)&is_not_None(object_2):
                                return object_,object_2 
                except:
                    pass
            if is_not_None(object_): #object_2 ищем только на листе с object_
                return object_,object_2
    return object_,object_2


# #### Задаем пути

# In[ ]:


#ячейка парсит все файлы в folder
root_path_parsing="Data for parsing/"
path_parsing="2017/2017_12"
full_path_parsing=os.path.join(root_path_parsing,path_parsing)

root_path_save="Preprocessed files"
full_path_save=os.path.join(root_path_save,path_parsing)
# full_path_save=os.path.join(root_path_save,"2015 preprocessed")

#Создаем директорию для сохранения файлов
isExist = os.path.exists(full_path_save)
if not isExist:
    os.makedirs(full_path_save)

# # #Пробую переименовать все файлы в директории с xlsx к xls
# transform_all_xlsx_files_to_xls(full_path_parsing)


# In[ ]:


'''
    For the given path, get the List of all files in the directory tree 
'''
def getListOfFiles(dirName):
    # create a list of file and sub directories 
    # names in the given directory 
    listOfFile = os.listdir(dirName)
    allFiles = list()
    # Iterate over all the entries
    for entry in listOfFile:
        # Create full path
        fullPath = os.path.join(dirName, entry)
        # If entry is a directory then get the list of files in this directory 
        if os.path.isdir(fullPath):
            allFiles = allFiles + getListOfFiles(fullPath)
        else:
            allFiles.append(fullPath)
                
    return allFiles


# row_values_not_none ['Подготовительные работы', 'план']

# In[ ]:


not_preprocessed_file=[]

# #Сам процесс парсинга
# for folder in tqdm_notebook(os.listdir(full_path_parsing)):
#генерируем все возможные пути папок
all_paths=getListOfFiles(full_path_parsing)

for full_path in tqdm_notebook(all_paths):
    file=full_path.split("\\")[-1]
    last_folder="\\".join(full_path.split("\\")[:-1]) #Нужна финальная директория, без файла

    list_for_file_saved=False #маркер успешного сохранения минимум 1 листа

    filename, file_extension = os.path.splitext(file)
    if (file_extension not in [".db",".doc",".docx"])&("thumbs" not in filename.lower()):

        print(f"Началась обработка файла {file}")
#             full_path=os.path.join(last_folder,file)
        month_for_parsing=get_month(filename)

        #Часть файлов в xlsx, их сначала переводим в xls, далее 
        try:
            book = xlrd.open_workbook(full_path)
            type_file="xls"
        except Exception as e:
            if str(e) == "Workbook is encrypted":
                book = decrypt_xlsx_files(file,last_folder) #Расшифровка зашифрованного файла
                type_file="xls"
            else:
                try:
                    book = openpyxl.load_workbook(full_path,data_only=True) #! data_only сохраняет последнее значение ,
                    # добавил
                    
                    type_file="xlsx"
                except:
                    try:
                        book = openpyxl.load_workbook(full_path,read_only=True,data_only=True) #! data_only сохраняет последнее значение ,
                        #read_only=True добавил
                        
                        type_file="xlsx"
                    except:
                        rename_path=full_path.replace(file_extension,".xlsx")
                        os.rename(full_path,rename_path)

                        book = openpyxl.load_workbook(rename_path,data_only=True) #! data_only сохраняет последнее значение
                        type_file="xlsx"

        #Задаем книгу для расшифровки дат
        def date_preproc_1(date_val,book=book):
            return date_preproc(date_val,book)

        if type_file=="xls":
            sheets=book.sheet_names()
        if type_file=="xlsx":
            sheets=book.sheetnames

        for sh_name in sheets:
            if (re.search(month_for_parsing,sh_name.lower())!=None)|("Лот"in filename): #not is_file_already_preprocessed(full_path,sh_name)
                if re.search("правила",str(sh_name).lower())==None: #Не интересуют листы правила заполнения/составления МСГ
                    #Файл для сохранения ресурсов
                    file_sheet_json={
                                    'file_name': file.strip(".xls").strip(".xlsx")+"_"+sh_name,
                                    'object':get_object(book,sheets,month_for_parsing,type_file)[0],
                                    'object_1':get_object(book,sheets,month_for_parsing,type_file)[1],
                                    'work': [],
                                    'resource': [],

                                }

                    print(f"Началась обработка листа {sh_name}")

                    if type_file=="xls":
                        sh = book.sheet_by_name(sh_name)
                    if type_file=="xlsx":
                        sh = book[sh_name]
                    print("Тип файла определен как",type_file)
                    #ОБРАБОТКА ЛИСТА
                    #Ищем нужные строки, ограничивающие таблицу - первая и последняя строка таблицы, последняя строка заголовков
                    start_table_rx=None
                    end_table_rx=None
                    last_head_rx=None
                    start_next_table_rx=None


                    if type_file=="xls":
                        num_rows = sh.nrows
                    if type_file=="xlsx":
                        num_rows = sh.max_row
    #                 print("num_rows",num_rows)

                    #Для xls перебираем от 0 до  num_rows-1, который последний, т.е. перебираем все
                    #Для xlsx перебираем от 1 до num_rows , который последний, т.е. перебираем все
                    print("num_rows",num_rows)
                    for rx in range(num_rows):
#                         print(rx)
                        if type_file=="xls":
                            row=sh.row(rx)

                        if type_file=="xlsx":
                            rx+=1 #Начинается с единицы
                            row=sh[rx]



#                         print("rx",rx)
                        row_values=[
                            re.sub("\n"," ",str(cell.value)).strip() for cell in row if (cell.value!="")&(cell.value!=None)]
    #                     print("finding borders table",row_values)
                        row_line="|".join(row_values)
#                             print("row_line",row_line)


                        if len(row_values)!=0: #Пустые строки не интересны для поиска границ таблиц

                            first_val=row_values[0]

                            #Проверка наличия маркера старта таблицы
                            start_table_markers=["п/п","наименование должностей, профессий","наименование и марка техники"] 
                            marker_in_line=lambda line: any([(m in str(line).lower()) for m in start_table_markers])  

    #                         print(rx,num_rows-1,type(rx))

                            #Ищем границы таблицы
                            if (any([marker_in_line(val) for val in row_values]))&(start_table_rx==None):
                                start_table_rx=rx
    #                             print("start_table_rx определено start_table_marker")
                            elif (any([marker_in_line(val) for val in row_values]))&(start_table_rx!=None):
                                end_table_rx=rx-1
                                start_next_table_rx=rx
    #                             print("end_table_rx определено start_table_marker")


                            elif (rx==num_rows-1)|(re.search("Процент выполнения МСГ",row_line)!=None): #уязвимость возможно
                                end_table_rx=rx
    #                             print("end_table_rx определено num_rows-1")
                            elif (re.search("ИТОГО",row_line)!=None):
                                end_table_rx=rx+1 

                            elif (sum([1 for v in list(range(1,29)) if str(int(v)) in row_line])>=26)&(last_head_rx==None): #должно быть минимум 25 дат
                                last_head_rx=rx #!
    #                             print("last_head_rx определено")

                            else:
                                pass
                        if rx==num_rows-1:
                            end_table_rx=rx
                            print("end_table_rx определено num_rows-1")

                        if all(v is not None for v in [start_table_rx,end_table_rx,last_head_rx]):
                            print(start_table_rx,end_table_rx,last_head_rx)
    #                         try:
                            table_to_json(file_sheet_json,
                                          sh,
                                          start_table_rx,
                                          end_table_rx,
                                          last_head_rx,
                                          type_file
                                         )

    #                         except:
    #                             print(
    #                 f"Не удалось сохранить таблицу в json, {file} {sh_name} {(start_table_rx,end_table_rx,last_head_rx)}"
    #                             )
    #                             #Показываю ошибку
    #                             traceback.print_exc()

                            #Обновляем индексы    
                            start_table_rx=None
                            end_table_rx=None
                            last_head_rx=None

                            if start_next_table_rx!=None:
                                start_table_rx=start_next_table_rx
                                start_next_table_rx=None

                                print("Индексы обновлены",start_table_rx,end_table_rx,last_head_rx)

                        #Если в строке ПОДПИСИ СТОРОН - она финальная, дальше смотреть строки не нужно 
                        if re.search("ПОДПИСИ СТОРОН",row_line)!=None: 
                            break

    #                 print(start_table_rx,end_table_rx,last_head_rx)
                    ###ЛИСТ ПРОЧИТАЛИ СОХРАНЯЕМ РЕЗУЛЬТАТ

                    #Saving json
                    with open(os.path.join(full_path_save,create_file_name(filename,
                                                                           path_parsing,
                                                                           full_path_save)+".json"),
                              "w",
                              encoding='utf8') as save_file:

                        json.dump(file_sheet_json,save_file)
                    list_for_file_saved=True
    #                 #Save preprocessed file name
    #                 append_preprocessed_file(full_path,sh_name)

            else:
                print(f"Не обработан {full_path} {sh_name}")
        if not list_for_file_saved:
            not_preprocessed_file.append(filename)
    else:
        #Если формат не подходит - нужно добавить файл в необработанные для правильного подсчета обработанных файлов
        not_preprocessed_file.append(filename)
        
not_preprocessed_file


# In[ ]:


not_preprocessed_file


# In[ ]:


len(not_preprocessed_file)+159


# In[ ]:


file="83_ТИП_xx.20.2017.json"
path=os.path.join(full_path_save,file)
print(path)
with open(path,"r",encoding="utf-8") as read_file:
    dat=json.load(read_file)
dat


# Заметки
