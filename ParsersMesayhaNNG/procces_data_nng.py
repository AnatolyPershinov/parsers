

import openpyxl
import pandas as pd
from pyxlsb import open_workbook as open_xlsb
from datetime import datetime
import os
from datetime import timedelta
import xlrd
from pyxlsb import open_workbook
import copy
M = {'Январь':1, 'Февраль':2,'Март':3, 'Апрель':4, 'Май':5, 'Июнь':6, 'Июль':7, 'Август':8, 'Сентябрь':9, 'Октябрь':10, 'Ноябрь':11, 'Декабрь':12}

def create_data(m,y):
    m_start = m 
    date = datetime(y, m, 1)#.date()
    d = timedelta(days=1)
    Dates = [date]
    while m_start == m:
        date+=d
        m_start = date.month
        if m_start==m:
            Dates.append(date)
    del m_start,date,d
    return Dates

def del_UPData(Data,index_names):
    delData=Data[0:index_names]
    if len(delData)!=0:
        for d in delData:
            Data.remove(d)
    return Data

def del_EmptyLines(Data):
    l = [None] * len(Data[0])
    delData=[]
    for d in Data:
        if d==l:
            delData.append(d)
    if len(delData)!=0:
        for d in delData:
            Data.remove(d)
    return Data


def delIdWeeksDaysMonth(Data,index_works):
    indexS,index_plan = None,None
    Keys=[]
    for i in range(len(Data)):
        if Data[i][index_works] in ['СВОД','СВОД за месяц','График производства работ','ИТОГО','Итого']:#
            indexS = i
            break  
          
    planKeys = Data[0] #   ПРЕДПАЛАГАЕМ ЧТО СТОЛБЕЦ ПЛАН ФАКТ УКАЗАН В 1 СТРОКЕ
    
    if 'план' in planKeys:
        index_plan = planKeys.index('план')# #['План''план/факт''План/Факт''дни месяца''Дни месяца']
    elif 'Дни месяца' in planKeys:
        index_plan = planKeys.index('Дни месяца')
    elif 'План' in planKeys:
        index_plan = planKeys.index('План')
    elif 'план / факт' in planKeys:
        index_plan = planKeys.index('план / факт')
    elif 'план/факт' in planKeys:
        index_plan = planKeys.index('план/факт')
    elif 'План/Факт' in planKeys:
        index_plan = planKeys.index('План/Факт')
#    elif ''
##        print(index_plan)
    del planKeys
    
    if indexS!=None:
        id_data = 1
        Kk,delData=[],[]
        for i in range(indexS):
            delData.append(Data[i])
            if i==0:
#                for j in range(index_plan+1, len(Data[i])):
#                    Data[i][j] = None
                Kk.append(Data[i])
            if i==id_data:# Предпалагаем, что 1 СТРОКА СОДЕРЖИТ ДАТЫ ПЛАНОВ 
                D = Data[i]
                                
        for k in delData:
            Data.remove(k)
        del i
        
        for j in range(index_plan):
            if Kk[0][j]==None:
                Kk[0][j] = Kk[0][j-1]
        Keys=Kk[0]
        
        for i in range(index_plan):
            if D[i]!=None:
                Keys[i]=str(Keys[i])+'_'+str(D[i])
        
        
        for i in range(index_plan+1,len(Keys)):
            Keys[i]=D[i]
    else:
        id_data = 1
        D = Data[id_data]
        
        for j in range(index_plan):
            if Data[0][j]==None:
                Data[0][j] = Data[0][j-1]
        Keys=Data[0]
        for i in range(index_plan):
            if D[i]!=None:
                Keys[i]=Keys[i]+'_'+str(D[i])
        for i in range(index_plan+1,len(Keys)):
            Keys[i]=Data[id_data][i]
            
        Data.remove(Data[id_data])
        Data.remove(Data[0])
        Data.remove(Data[0])

    
    return Data,index_plan,Keys,indexS
 
def lastFact(Data,index_plan):
    lastF = None
    for i in range(len(Data)):
        if Data[i][index_plan]=='Факт':
            lastF = i
    if lastF!=None:
        delData = Data[lastF+1:]
        for d in delData:
            Data.remove(d)
    return Data    

def del_rows(Data,Keys, indexPlan):
    indexDate = indexPlan + 1 #Предполагаем, ЧТО ГРАФИКИ НАЧИНАЮТСЯ В СЛЕД СТОЛБЦЕ ОТ ПЛАН/ФАКТ
    print(indexDate,type(Keys[indexDate]))
    if type(Keys[indexDate])!= datetime:
        first_data_in_data = xlrd.xldate.xldate_as_datetime(Keys[indexDate], 0).date()
        print(first_data_in_data, Dates[0])
        if first_data_in_data!=Dates[0]:
            for i in range(indexDate, len(Keys)):
                newdate = xlrd.xldate.xldate_as_datetime(Keys[i], 0).date()
                if newdate==Dates[0]:
                    delLine = i
                    break
            for  i in range(indexDate, len(Keys)):
                newdate = xlrd.xldate.xldate_as_datetime(Keys[i], 0).date()
                if newdate==Dates[-1]:
                    indexLastDate = i
                    break
            bufData = []
            for data in Data:
                bufData.append(data[0:indexDate]+data[delLine:indexLastDate+1])
            Data = bufData
            Keys = Keys[0:indexDate]+Dates#Keys[delLine:indexLastDate+1]       #
        else:
            delLine = indexDate
            indexLastDate = delLine + len(Dates)-1
            bufData = []
            for data in Data:
                bufData.append(data[0:indexDate]+data[delLine:indexLastDate+1])
            Data = bufData
            Keys = Keys[0:indexDate]+Dates#Keys[delLine:indexLastDate+1]       #

    else:
        first_data_in_data = Keys[indexDate]#.date()
        print(first_data_in_data, Dates[0])
        if first_data_in_data!=Dates[0]:
            for i in range(indexDate,len(Keys)):
                
                try:
                    if Keys[i].date()==Dates[0]:
                        delLine = i
                        break     
                except: pass
            for  i in range(indexDate, len(Keys)):
                try:
                    if Keys[i].date()==Dates[-1]:#2
                        indexLastDate = i
                        break
                except:pass
            bufData = []            
            for data in Data:
                bufData.append(data[0:indexDate]+data[delLine:indexLastDate+1])
            Data = bufData
            Keys = Keys[0:indexDate]+Keys[delLine:indexLastDate+1]
        else:
            delLine = 10
            indexLastDate=37
            for i in range(indexDate,len(Keys)):
                print(Keys[i],Dates[0])
                
                if Keys[i].date()==Dates[0]:
                    delLine = i
                    break
            for  i in range(indexDate, len(Keys)):
                if Keys[i].date()==Dates[-1]:
                    indexLastDate = i
                    break
            bufData = []            
            for data in Data:
                bufData.append(data[0:indexDate]+data[delLine:indexLastDate+1])                        
            Data = bufData
            Keys = Keys[0:indexDate]+Keys[delLine:indexLastDate+1]
    return Data,Keys,indexDate

def metadata(Data,index_works,indexPlan):
    Prop = []
    for i in range(len(Data)):
        newList = Data[i][index_works+1:indexPlan-1] 
        l = [None] * len(newList)
        if newList == l:
            Prop.append((i,Data[i][index_works],Data[i]))           
    Pp = copy.deepcopy(Prop)
    num_prop=0
    while len(Pp)!=0:
        l=[]
        for i in range(len(Pp)):   
            if Pp[i][0]- Pp[0][0] == i:
                l.append(Pp[i])
        if len(l)>num_prop:
            num_prop = len(l)
        for l0 in l:
            Pp.remove(l0)
#    del l,i
    P = {}
    for i in range(num_prop):
        P['property '+ str(i+1)] =[]
    while len(Prop)!=0:
        l = []
        for i in range(len(Prop)):   
            if Prop[i][0]- Prop[0][0] == i:
                l.append(Prop[i])
        if len(l) == num_prop:
            for i in range(len(l)):
                P['property '+ str(i+1)].append(l[i])
        else:
            num = num_prop-len(l)+1
            for i in range(len(l)):
                P['property '+ str(num+i)].append(l[i])
        for l0 in l:
            Prop.remove(l0)  
    return P

def full_metadata(Data,P,indexPlan):    
    for key in P:
        if len(P[key])!=0:
            for i in range(len(P[key])-1):
                start,finish = P[key][i][0], P[key][i+1][0]
                for j in range(start,finish):
                    Data[j].append(P[key][i][1])
            last = P[key][-1]
            start,finish = last[0], len(Data)
            for j in range(start,finish):
                Data[j].append(last[1])
            Keys.append(key)
    for key in P:
        for p in P[key]:
            Data.remove(p[2])
    del key,p
    for d in Data:
        if len(d)!=len(Keys):
            nn = len(Keys)-len(d)
            for i in range(nn):
                d.append(None)
    return Data



m = 'Февраль'
y = 2021
path = 'СМГ ННГ/Проект Газ '+str(y)+'/'

for file in os.listdir(path):
    if '_processed' not in file and '_ресурсы' not in file:
        filename, file_extension = os.path.splitext(file)
        print(file)
        Data=[]
        i=0
        if file_extension == '.xlsb':       
            workbook = open_workbook(path+file)
            with open_xlsb(path+file) as wb:
                print(wb.sheets)
                for sheet_name in wb.sheets:                    
                    if sheet_name in ['МСГ','СМГ','СтройПроекСервис']: 
                        finalName = filename+'_'+sheet_name+'_processed'
                        with wb.get_sheet(sheet_name) as sheet: 
                            for row in sheet.rows():
                                values = [r.v for r in row] # 
                                if 'Наименование работ' in values:
                                    index_names = i
                                    index_works = values.index('Наименование работ')
                                elif 'Наименование' in values :
                                    index_names = i
                                    index_works = values.index('Наименование')
                                    values[index_works] = 'Наименование работ'
                                elif '     Наименование работ' in values:
                                    index_names = i
                                    index_works = values.index('     Наименование работ')
                                    values[index_works] = 'Наименование работ'                                    
                                Data.append(values)       
                                i+=1   
                                del values                                
        elif file_extension == '.xlsx' or file_extension == '.xlsm':
            book = openpyxl.load_workbook(path+file,data_only=True)
            print(book.sheetnames)
            for sheet_name in book.sheetnames:
                if sheet_name in ['УКПГ 2020-06','Лист2' ]:#'куст1','куст 4'
                    finalName = filename+'_'+sheet_name+'_processed'
                    worksheet = book[sheet_name]
                    for row in worksheet.rows:
                        values = [r.value for r in row] 
                        if 'Наименование работ' in values:
                            index_names = i
                            index_works = values.index('Наименование работ')
                        elif '     Наименование работ' in values:
                            index_names = i
                            index_works = values.index('     Наименование работ') #index_works = values.index('Наименование работ')
                            values[index_works] = 'Наименование работ'
                        elif 'Название работы' in values:
                            index_names = i
                            index_works = values.index('Название работы')
                            values[index_works] = 'Наименование работ'
                        elif 'Название задачи' in values:
                            index_names = i    
                            index_works = values.index('Название задачи')
                            values[index_works] = 'Наименование работ'
                        elif 'Виды работ' in values:
                            index_names = i    
                            index_works = values.index('Виды работ')
                            values[index_works] = 'Наименование работ'
                        elif 'Наименование видов и/или этапов работ' in values:
                            index_names = i    
                            index_works = values.index('Наименование видов и/или этапов работ')
                            values[index_works] = 'Наименование работ'
                            
                            
                        Data.append(values)        
                        i+=1   
                        del values
                    del row
            
            del i,sheet_name
        print('СТРОК ВСЕГО:',len(Data))  
        
        if len(Data)!=0:    
            Dates=create_data(M[m],y)
            
            Data = del_UPData(Data,index_names)
            del index_names
            print('СТРОК 1 чистка:',len(Data))
            Data = del_EmptyLines(Data)
            print('СТРОК 2 чистка:',len(Data))
###            
            Data,indexPlan,Keys,indexS = delIdWeeksDaysMonth(Data,index_works)

            delData = []
            for i in range(len(Keys)):
                if type(Keys[i])==str and ('ИТОГО' in Keys[i] or 'неделя' in Keys[i]):
                    print(i,Keys[i])
                    delData.append(i)
            delData.reverse()
            for i in delData:
                Keys.pop(i)
                for d in Data:
                    d.pop(i)
            
            K = Keys[:10]+ Keys[256:284]
            for i in range(len(Dates)):
                K[indexPlan+1+i]=Dates[i]
            newData = []
            for d in Data:
                nw = d[0:10]+d[256:284]
                newData.append(nw)
##            
            Data,Keys,indexDate = del_rows(newData,K, indexPlan)

            Data = lastFact(Data,indexPlan)    
#             #####  FILL DATA   ###########
            for i in range(len(Data)):
                if Data[i][indexPlan]  == 'факт':
                    for j in range(len(Data[i])):
                        if j<indexPlan:
                            if Data[i][j]==None:
                                Data[i][j] = Data[i-1][j]
                    del j
                del i    
            P = metadata(Data,index_works,indexPlan)
            Data = full_metadata(Data,P,indexPlan)

###################################################################
            for i in range(10,38):
                print(i)
                newdate = Keys[i].date()
                Keys[i]=newdate
            PrData = {}
            for i in range(len(Keys)):
                    key = Keys[i]
                    D=[]
                    for d in Data:
                        D.append(d[i])
                    PrData[key] = D
            del i,d,D  
            
            z = pd.DataFrame(PrData)
            with pd.ExcelWriter(path + finalName+".xlsx", engine='xlsxwriter') as writer:  
                z.to_excel(writer, sheet_name='МСГ', index=False)            
###################################################################
#
#
#            
            
            
            
            
            
            
            
            
            
            
            
            
            
            
            
            
            
            
            
            
            